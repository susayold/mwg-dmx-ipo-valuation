from __future__ import annotations

import calendar
import csv
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from .data_io import file_sha256


ALLOWED_SHEETS: tuple[str, ...] = (
    "Balance Sheet",
    "Income Statement",
    "Cash Flow Statement",
)

STATEMENT_CODES = {
    "Balance Sheet": "balance_sheet",
    "Income Statement": "income_statement",
    "Cash Flow Statement": "cash_flow_statement",
}

KEY_LINE_ITEM_MAP: dict[str, dict[str, str]] = {
    "balance_sheet": {
        "A. CURRENT ASSETS": "current_assets",
        "I. Cash & cash equivalents": "cash_and_cash_equivalents",
        "III. Current accounts receivable": "current_receivables",
        "IV. Inventories": "inventory",
        "B. NON-CURRENT ASSETS": "non_current_assets",
        "TOTAL ASSETS": "total_assets",
        "C. LIABILITIES": "total_liabilities",
        "I. Current liabilities": "current_liabilities",
        "8. Short-term loans": "short_term_loans",
        "D. EQUITY": "total_equity",
        "TOTAL LIABILITIES & EQUITY": "total_liabilities_and_equity",
    },
    "income_statement": {
        "Net revenue": "revenue",
        "Cost of goods sold": "cogs",
        "Gross profit": "gross_profit",
        "- Interest expense": "interest_expense",
        "Operating profit": "operating_profit",
        "Profit before tax": "pretax_income",
        "Net profit after tax": "net_income",
        "NPAT attributable to shareholders of the parent": "net_income_attributable_parent",
        "NPAT attributable to non-controlling interests": "net_income_attributable_nci",
        "Basic earnings per share": "basic_eps_vnd",
        "Diluted earnings per share": "diluted_eps_vnd",
    },
    "cash_flow_statement": {
        "Profit before tax": "pretax_income_cash_flow",
        "Depreciation & amortisation of fixed assets": "depreciation_and_amortisation",
        "Interest expense": "interest_expense_cash_flow",
        "Operating profit before WC changes": "operating_profit_before_working_capital",
        "Net cash from operating activities": "operating_cash_flow",
        "Purchase and construction of fixed asset": "capex",
        "Net cash from (used in) investing activities": "investing_cash_flow",
        "Net cash from used in financing activities": "financing_cash_flow",
        "Net change in cash": "net_change_in_cash",
        "Cash and cash equipvalents at beginning of period": "opening_cash",
        "Impact of exchange rate fluctuation": "fx_cash_effect",
        "Cash and cash equipvalents at end of period": "ending_cash",
    },
}


class WorkbookFormatError(ValueError):
    """Raised when an allowlisted workbook sheet does not match the contract."""


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    normalized = normalized.lower().replace("&", " and ")
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    if normalized and normalized[0].isdigit():
        normalized = "item_" + normalized
    return normalized or "unnamed_line_item"


def _period_end(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip().upper().replace(" ", "")
    match = re.fullmatch(r"([1-4])Q(\d{4})", text) or re.fullmatch(r"Q([1-4])(\d{4})", text)
    if match:
        quarter = int(match.group(1))
        year = int(match.group(2))
        month = quarter * 3
        day = calendar.monthrange(year, month)[1]
        return date(year, month, day).isoformat()
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise WorkbookFormatError(f"Unsupported period header {value!r}") from exc


def _whole_vnd(value: Any) -> int | None:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return None
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _sheet_rows(worksheet: Any) -> list[tuple[Any, ...]]:
    # Some published workbooks declare an A1 dimension despite containing data.
    # reset_dimensions lets openpyxl stream the actual allowlisted worksheet.
    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.reset_dimensions()
    return [tuple(row) for row in worksheet.iter_rows(values_only=True)]


def _extract_sheet(worksheet: Any, statement: str, source_id: str) -> list[dict[str, Any]]:
    rows = _sheet_rows(worksheet)
    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if len(row) >= 2
            and str(row[0] or "").strip().casefold() in {"chỉ tiêu", "chi tieu"}
            and str(row[1] or "").strip().casefold() == "items"
        ),
        None,
    )
    if header_index is None:
        raise WorkbookFormatError(f"{worksheet.title}: cannot locate Items header")

    unit_declared_vnd = any(
        len(row) > 1 and str(row[1] or "").strip().casefold() == "unit: vnd"
        for row in rows[:header_index]
    )
    if not unit_declared_vnd:
        raise WorkbookFormatError(
            f"{worksheet.title}: expected an explicit 'Unit: VND' declaration"
        )

    header = rows[header_index]
    periods: dict[int, str] = {}
    for column_index, value in enumerate(header[2:], start=2):
        if value is not None and str(value).strip():
            periods[column_index] = _period_end(value)
    if not periods:
        raise WorkbookFormatError(f"{worksheet.title}: no period columns found")

    row_specs: list[tuple[int, str, tuple[Any, ...]]] = []
    for excel_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        english_label = str(row[1] or "").strip() if len(row) > 1 else ""
        if not english_label:
            continue
        if not any(
            column_index < len(row) and _whole_vnd(row[column_index]) is not None
            for column_index in periods
        ):
            continue
        mapped = KEY_LINE_ITEM_MAP.get(statement, {}).get(english_label)
        base_line_item = mapped or _slugify(english_label)
        row_specs.append((excel_row, base_line_item, row))

    counts = Counter(base_line_item for _, base_line_item, _ in row_specs)
    output: list[dict[str, Any]] = []
    for excel_row, base_line_item, row in row_specs:
        line_item = (
            f"{base_line_item}__row_{excel_row}"
            if counts[base_line_item] > 1
            else base_line_item
        )
        for column_index, period in periods.items():
            if column_index >= len(row):
                continue
            value_vnd = _whole_vnd(row[column_index])
            if value_vnd is None:
                continue
            output.append(
                {
                    "statement": statement,
                    "line_item": line_item,
                    "period": period,
                    "value_vnd": value_vnd,
                    "source_id": source_id,
                }
            )
    return output


def validate_extracted_rows(rows: Iterable[dict[str, Any]]) -> None:
    """Enforce key uniqueness and the three primary accounting bridges."""

    rows = list(rows)
    required_fields = {"statement", "line_item", "period", "value_vnd", "source_id"}
    if not rows:
        raise WorkbookFormatError("Extracted dataset is empty")
    for index, row in enumerate(rows, start=1):
        if set(row) != required_fields:
            raise WorkbookFormatError(f"Extracted row {index} does not match the CSV contract")
        if not row["source_id"]:
            raise WorkbookFormatError(f"Extracted row {index} has no source_id")

    keys = [
        (str(row["statement"]), str(row["line_item"]), str(row["period"]))
        for row in rows
    ]
    duplicates = [key for key, count in Counter(keys).items() if count > 1]
    if duplicates:
        raise WorkbookFormatError(f"Extracted facts contain duplicate keys: {duplicates[:3]}")

    values = {
        (str(row["statement"]), str(row["line_item"]), str(row["period"])): int(
            row["value_vnd"]
        )
        for row in rows
    }
    periods = sorted({str(row["period"]) for row in rows})
    for period in periods:
        balance_keys = [
            ("balance_sheet", item, period)
            for item in ("total_assets", "total_liabilities", "total_equity")
        ]
        if all(key in values for key in balance_keys):
            assets, liabilities, equity = (values[key] for key in balance_keys)
            if assets != liabilities + equity:
                raise WorkbookFormatError(
                    f"Balance sheet does not reconcile for {period}: "
                    f"difference={assets - liabilities - equity} VND"
                )

        income_keys = [
            ("income_statement", item, period)
            for item in ("revenue", "cogs", "gross_profit")
        ]
        if all(key in values for key in income_keys):
            revenue, cogs, gross_profit = (values[key] for key in income_keys)
            if revenue + cogs != gross_profit:
                raise WorkbookFormatError(
                    f"Gross profit does not reconcile for {period}: "
                    f"difference={revenue + cogs - gross_profit} VND"
                )

        cash_keys = [
            ("cash_flow_statement", item, period)
            for item in ("opening_cash", "net_change_in_cash", "fx_cash_effect", "ending_cash")
        ]
        if all(key in values for key in cash_keys):
            opening, change, fx_effect, ending = (values[key] for key in cash_keys)
            if opening + change + fx_effect != ending:
                raise WorkbookFormatError(
                    f"Cash bridge does not reconcile for {period}: "
                    f"difference={opening + change + fx_effect - ending} VND"
                )


def extract_workbook_rows(workbook_path: str | Path) -> list[dict[str, Any]]:
    """Extract only the three allowlisted statement sheets.

    Sheet names may be listed by the workbook package, but cell contents from
    `config`, `notes`, or any other tab are never accessed.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "The DMX adapter requires openpyxl; install with pip install -e '.[xlsx]'"
        ) from exc

    path = Path(workbook_path)
    digest = file_sha256(path)
    source_id = f"DMX_Q1_2026_DATA_PACK_XLSX_{digest.upper()}"
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        missing = [sheet for sheet in ALLOWED_SHEETS if sheet not in workbook.sheetnames]
        if missing:
            raise WorkbookFormatError(f"Workbook is missing allowlisted sheets: {missing}")
        output: list[dict[str, Any]] = []
        for sheet_name in ALLOWED_SHEETS:
            statement = STATEMENT_CODES[sheet_name]
            output.extend(_extract_sheet(workbook[sheet_name], statement, source_id))
        validate_extracted_rows(output)
        return output
    finally:
        workbook.close()


def write_extracted_csv(rows: Iterable[dict[str, Any]], output_path: str | Path) -> int:
    rows = list(rows)
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["statement", "line_item", "period", "value_vnd", "source_id"]
    with target.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def extract_to_csv(workbook_path: str | Path, output_path: str | Path) -> int:
    return write_extracted_csv(extract_workbook_rows(workbook_path), output_path)
