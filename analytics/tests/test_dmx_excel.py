from __future__ import annotations

import csv
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

ANALYTICS_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ANALYTICS_ROOT / "src"))

from mwg_dmx_analytics.dmx_excel import (
    ALLOWED_SHEETS,
    WorkbookFormatError,
    extract_to_csv,
    extract_workbook_rows,
    validate_extracted_rows,
)

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None


@unittest.skipIf(Workbook is None, "openpyxl optional dependency is not installed")
class DmxExcelAdapterTests(unittest.TestCase):
    def _fixture_workbook(self, path: Path) -> None:
        workbook = Workbook()
        workbook.active.title = "Balance Sheet"
        workbook.create_sheet("Income Statement")
        workbook.create_sheet("Cash Flow Statement")
        workbook.create_sheet("config")
        workbook.create_sheet("notes")

        for sheet_name in ALLOWED_SHEETS:
            sheet = workbook[sheet_name]
            sheet.append(["Entity", "Entity English"])
            sheet.append(["Statement", sheet_name])
            sheet.append(["Đơn vị tính: đồng", "Unit: VND"])
            if sheet_name == "Balance Sheet":
                sheet.append(["Chỉ tiêu", "Items", datetime(2026, 3, 31), datetime(2025, 12, 31)])
                sheet.append(["Tổng tài sản", "TOTAL ASSETS", 100.49, 90])
            elif sheet_name == "Income Statement":
                sheet.append(["Chỉ tiêu", "Items", "1Q2026", "1Q2025"])
                sheet.append(["Doanh thu", "Net revenue", 50, 40])
            else:
                sheet.append(["Chỉ tiêu", "Items", "1Q2026", "1Q2025"])
                sheet.append(["LCTT HĐKD", "Net cash from operating activities", 7, 6])

        workbook["config"]["A1"] = "DO_NOT_EXTRACT_CONFIG_SECRET"
        workbook["notes"]["A1"] = "DO_NOT_EXTRACT_NOTES_SECRET"
        workbook.save(path)

    def test_only_allowlisted_statement_cells_are_extracted(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "fixture.xlsx"
            self._fixture_workbook(workbook_path)
            rows = extract_workbook_rows(workbook_path)
            self.assertEqual(len(rows), 6)
            self.assertEqual(
                {row["statement"] for row in rows},
                {"balance_sheet", "income_statement", "cash_flow_statement"},
            )
            serialized = repr(rows)
            self.assertNotIn("DO_NOT_EXTRACT", serialized)
            self.assertEqual(rows[0]["value_vnd"], 100)
            self.assertTrue(rows[0]["source_id"].startswith("DMX_Q1_2026_DATA_PACK_XLSX_"))

    def test_csv_contract_is_exact(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "fixture.xlsx"
            output_path = Path(directory) / "facts.csv"
            self._fixture_workbook(workbook_path)
            row_count = extract_to_csv(workbook_path, output_path)
            with output_path.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.DictReader(handle)
                rows = list(reader)
            self.assertEqual(row_count, 6)
            self.assertEqual(
                reader.fieldnames,
                ["statement", "line_item", "period", "value_vnd", "source_id"],
            )
            self.assertEqual(rows[0]["period"], "2026-03-31")

    def test_duplicate_extracted_key_is_rejected(self) -> None:
        row = {
            "statement": "income_statement",
            "line_item": "revenue",
            "period": "2026-03-31",
            "value_vnd": 10,
            "source_id": "SOURCE",
        }
        with self.assertRaises(WorkbookFormatError):
            validate_extracted_rows([row, row.copy()])


if __name__ == "__main__":
    unittest.main()
