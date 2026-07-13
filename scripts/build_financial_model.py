"""Build the MWG/DMX IPO and SOTP portfolio model.

The workbook deliberately separates official facts, company outlook and
illustrative analyst assumptions.  It is a portfolio case study, not investment
advice.  Run from the repository root with::

    python scripts/build_financial_model.py

The generated workbook is written to ``model/MWG_DMX_IPO_SOTP_Model.xlsx``.
No market conclusion is produced unless same-date price inputs are supplied.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = REPO_ROOT / "model" / "MWG_DMX_IPO_SOTP_Model.xlsx"
RAW_Q1_PATH = REPO_ROOT / "data" / "raw" / "dmx" / "dmx_q1_2026_data_pack.xlsx"

MODEL_VERSION = "1.0.0"
DATA_CUTOFF = date(2026, 7, 13)
# Excel stores naive datetimes; the repository timezone is Asia/Bangkok.
GENERATED_AT = datetime.now().replace(microsecond=0)

SHEET_ORDER = [
    "Cover",
    "Sources",
    "IPO_Bridge",
    "Q1_Actuals",
    "Mgmt_Forecast",
    "Assumptions",
    "DMX_Valuation",
    "MWG_SOTP",
    "Scenarios",
    "Checks",
]

# Palette inspired by institutional research templates.
NAVY = "0B1F3A"
NAVY_2 = "17365D"
TEAL = "00A6A6"
BLUE = "4472C4"
INPUT_BLUE = "0000FF"
LINK_GREEN = "008000"
WHITE = "FFFFFF"
BLACK = "202124"
GREY = "687078"
MID_GREY = "A6A6A6"
LIGHT_GREY = "F2F4F7"
LIGHT_BLUE = "D9EAF7"
LIGHT_TEAL = "DDEBF7"
PALE_GREEN = "E2F0D9"
PALE_YELLOW = "FFF2CC"
PALE_RED = "FCE4D6"
RED = "C00000"
ORANGE = "ED7D31"

THIN_GREY = Side(style="thin", color="D9DEE5")
MEDIUM_NAVY = Side(style="medium", color=NAVY)

FMT_BN = '#,##0.0;[Red](#,##0.0);-'
FMT_BN_1 = '#,##0.0;[Red](#,##0.0);-'
FMT_BN_2 = '#,##0.00;[Red](#,##0.00);-'
FMT_SHARES = '#,##0.0000'
FMT_VND = '#,##0;[Red](#,##0);-'
FMT_PCT = '0.0%'
FMT_PCT_2 = '0.00%'
FMT_MULTIPLE = '0.0x'
FMT_DATE = 'dd-mmm-yyyy'


@dataclass(frozen=True)
class Source:
    source_id: str
    issuer: str
    document_type: str
    title: str
    period_end: date | None
    published_date: date | None
    retrieved_at: date
    audit_status: str
    scope: str
    url: str
    local_artifact: str
    use: str
    notes: str


SOURCES = [
    Source(
        "DMX-IPO-SALESKIT",
        "DMX",
        "Investor presentation / sales kit",
        "DMX IPO sales kit",
        date(2025, 12, 31),
        None,
        DATA_CUTOFF,
        "N/A",
        "Company presentation",
        "https://cdnv2-tmdt.tgdd.vn/mwgvn/investorrelations/files/posts/2026/5/5005818/d8/6b/d86b5d707c2c0b5baea9dbcffafaaa50.pdf",
        "data/raw/dmx/dmx_ipo_presentation_2026.pdf",
        "Planned IPO terms and management outlook",
        "Company-provided material; forecasts are not guarantees.",
    ),
    Source(
        "DMX-IPO-RES14",
        "DMX",
        "Board resolution / IPO result",
        "Resolution No. 14 on actual IPO result",
        None,
        date(2026, 6, 30),
        DATA_CUTOFF,
        "N/A",
        "Issuer disclosure",
        "https://cdnv2-tmdt.tgdd.vn/mwgvn/investorrelations/files/posts/2026/7/0/f1/08/f108ca2a7399d9a70674915a4b2cf651.pdf",
        "N/A (source linked; raw files are git-ignored)",
        "Actual allocated/cancelled/post-IPO shares and proceeds",
        "Primary source for the actual IPO bridge.",
    ),
    Source(
        "DMX-IPO-COMPLETE",
        "DMX",
        "Issuer announcement",
        "DMX completes IPO and raises more than VND 13,315bn",
        None,
        date(2026, 6, 30),
        DATA_CUTOFF,
        "N/A",
        "Issuer announcement",
        "https://www.dmx.vn/eng/news/dien-may-xanh-completes-landmark-ipo-raising-more-than-vnd-13-315-billion-and-lifting-charter-capital-to-vnd-12-677-billion-5002485",
        "N/A",
        "Rounded MWG ownership disclosure and transaction context",
        "'Nearly 86%' is a reasonableness check, not exact ownership.",
    ),
    Source(
        "DMX-Q1-2026",
        "DMX",
        "Interim financial statements and data pack",
        "DMX Q1 2026 consolidated financial statements",
        date(2026, 3, 31),
        None,
        DATA_CUTOFF,
        "Unaudited",
        "Consolidated",
        "https://cdnv2-tmdt.tgdd.vn/mwgvn/investorrelations/files/posts/2026/6/0/6b/e5/6be50a510e272451464dae07771401ca.xlsx",
        "data/raw/dmx/dmx_q1_2026_data_pack.xlsx",
        "Q1 2026 actual balance sheet, income statement and cash flow",
        "Model values are in VND bn; source workbook is in VND.",
    ),
    Source(
        "DMX-6M-2026",
        "DMX",
        "Business update",
        "DMX business results — six months 2026",
        date(2026, 6, 30),
        None,
        DATA_CUTOFF,
        "Unaudited",
        "Company LFL operating update",
        "https://cdnv2-tmdt.tgdd.vn/mwgvn/investorrelations/files/posts/2026/7/5006021/b5/ee/b5ee17c283b469b12b3b435d00da8aa4.pdf",
        "N/A",
        "Operating KPI context only",
        "Not used as a substitute for audited/interim financial statements.",
    ),
    Source(
        "DMX-OUTLOOK",
        "DMX",
        "Investor presentation",
        "DMX 2026–2030 management outlook",
        date(2030, 12, 31),
        None,
        DATA_CUTOFF,
        "N/A",
        "Company outlook",
        "https://cdnv2-tmdt.tgdd.vn/mwgvn/investorrelations/files/posts/2026/6/5005889/4a/f4/4af4a4152a74661a28063a23b4d6747f.pdf",
        "N/A",
        "Management revenue, NPAT and gross-margin outlook",
        "Separated from analyst forecast and actuals.",
    ),
    Source(
        "MWG-AR-2025",
        "MWG",
        "Annual report",
        "MWG Annual Report 2025",
        date(2025, 12, 31),
        None,
        DATA_CUTOFF,
        "Audited sections as identified in report",
        "Consolidated / corporate",
        "https://cdnv2.tgdd.vn/mwgvn/investorrelations/files/posts/2026/3/5677/48/8a/488ab752e1224bcbd9dd5413be04179f.pdf",
        "data/raw/mwg/mwg_annual_report_2025.pdf",
        "Group context and perimeter",
        "No current MWG market price is imported into this workbook.",
    ),
    Source(
        "MWG-ERABLUE-JV",
        "MWG",
        "Issuer article",
        "MWG describes EraBlue as a joint venture",
        None,
        None,
        DATA_CUTOFF,
        "N/A",
        "Issuer article",
        "https://mwg.vn/tin-tuc/erablue-buoc-chuyen-minh-cua-the-gioi-di-dong-o-quoc-gia-dong-dan-nhat-asean-179",
        "N/A",
        "Perimeter and double-counting control",
        "EraBlue is included once inside DMX in this illustrative model.",
    ),
]


# Official Q1 values embedded for a reproducible public build.  These values are
# exact VND bn conversions of the issuer's downloadable data pack.  When the raw
# workbook is present locally, the script verifies these constants against it.
Q1 = {
    "revenue_2026": 32_541.950091890,
    "revenue_2025": 25_153.558068114,
    "cogs_2026": -26_300.728779400,
    "cogs_2025": -20_627.386516636,
    "gross_profit_2026": 6_241.221312490,
    "gross_profit_2025": 4_526.171551478,
    "financial_income_2026": 512.117161615,
    "financial_income_2025": 436.151968360,
    "financial_expense_2026": -326.141821808,
    "financial_expense_2025": -202.423284684,
    "interest_expense_2026": -324.569874029,
    "interest_expense_2025": -201.641796928,
    "selling_expense_2026": -2_438.374451087,
    "selling_expense_2025": -2_139.579773103,
    "ga_expense_2026": -1_225.141003222,
    "ga_expense_2025": -771.423462048,
    "share_jv_2026": 9.011741928,
    "share_jv_2025": 3.129774300,
    "operating_profit_2026": 2_772.692939916,
    "operating_profit_2025": 1_852.026774303,
    "pbt_2026": 2_775.441460755,
    "pbt_2025": 1_856.564880868,
    "tax_2026": -556.874594506,
    "tax_2025": -378.189500224,
    "npat_2026": 2_218.566866249,
    "npat_2025": 1_478.375380644,
    "eps_2026": 2_015.0,
    "eps_2025": 1_342.0,
    "cash_mar_2026": 3_312.678561377,
    "cash_dec_2025": 3_578.154917657,
    "short_term_investments_mar_2026": 25_255.708750726,
    "short_term_investments_dec_2025": 25_246.515124040,
    "receivables_mar_2026": 1_612.050831121,
    "receivables_dec_2025": 1_432.564888990,
    "inventory_mar_2026": 23_054.446543337,
    "inventory_dec_2025": 22_759.101620225,
    "inventory_provision_mar_2026": -749.032356311,
    "inventory_provision_dec_2025": -609.594155009,
    "ppe_mar_2026": 751.596973808,
    "ppe_dec_2025": 840.472791325,
    "jv_investment_mar_2026": 502.561163711,
    "jv_investment_dec_2025": 371.159064528,
    "long_term_deposits_mar_2026": 1_499.502400000,
    "long_term_deposits_dec_2025": 1_400.000000000,
    "total_assets_mar_2026": 57_327.010309745,
    "total_assets_dec_2025": 56_916.237866410,
    "payables_mar_2026": 8_598.546695344,
    "payables_dec_2025": 9_701.594456166,
    "short_term_debt_mar_2026": 22_158.890775085,
    "short_term_debt_dec_2025": 23_429.114317650,
    "total_liabilities_mar_2026": 37_307.871832989,
    "total_liabilities_dec_2025": 39_115.666255903,
    "equity_mar_2026": 20_019.138476756,
    "equity_dec_2025": 17_800.571610507,
    "cfo_2026": 863.686766977,
    "cfo_2025": 2_477.049226791,
    "capex_2026": -116.154581487,
    "capex_2025": -24.589934831,
    "cfi_2026": 140.146492181,
    "cfi_2025": -1_936.035958785,
    "debt_drawdown_2026": 22_101.481775842,
    "debt_drawdown_2025": 17_652.457126709,
    "debt_repayment_2026": -23_371.705318407,
    "debt_repayment_2025": -19_751.444561031,
    "cff_2026": -1_270.223542565,
    "cff_2025": -2_598.987434326,
    "net_cash_change_2026": -266.390283407,
    "net_cash_change_2025": -2_057.974166320,
    "cash_open_2026": 3_578.154917657,
    "cash_open_2025": 3_780.128276727,
    "fx_effect_2026": 0.913927127,
    "fx_effect_2025": 0.048240903,
    "cash_close_2026": 3_312.678561377,
    "cash_close_2025": 1_722.202351310,
}


MGMT_FORECAST = {
    2025: {"revenue": 107_000.0, "npat": 6_075.0, "gross_margin": 0.171, "status": "Actual (company presentation)"},
    2026: {"revenue": 122_500.0, "npat": 7_350.0, "gross_margin": 0.175, "status": "Company outlook"},
    2027: {"revenue": 135_000.0, "npat": 8_500.0, "gross_margin": 0.178, "status": "Company outlook"},
    2028: {"revenue": 149_000.0, "npat": 9_800.0, "gross_margin": 0.180, "status": "Company outlook"},
    2029: {"revenue": 164_000.0, "npat": 11_300.0, "gross_margin": 0.183, "status": "Company outlook"},
    2030: {"revenue": 182_000.0, "npat": 13_000.0, "gross_margin": 0.184, "status": "Company outlook"},
}


def bn(vnd: float) -> float:
    return float(vnd) / 1_000_000_000


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def set_title(ws, title: str, subtitle: str | None = None, end_col: int = 11) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = fill(NAVY)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 31
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        cell = ws.cell(2, 1, subtitle)
        cell.fill = fill(NAVY_2)
        cell.font = Font(name="Aptos", size=10, italic=True, color="D9EAF7")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[2].height = 23


def section(ws, row: int, title: str, start_col: int = 1, end_col: int = 11) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, title)
    cell.fill = fill(NAVY_2)
    cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 21


def style_table_header(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill(TEAL)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN_GREY, bottom=THIN_GREY, left=THIN_GREY, right=THIN_GREY)
    ws.row_dimensions[row].height = 30


def style_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Aptos", size=9, color=BLACK)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN_GREY)


def style_input(cell, number_format: str | None = None) -> None:
    cell.fill = fill(PALE_YELLOW)
    cell.font = Font(name="Aptos", size=9, color=INPUT_BLUE)
    cell.protection = Protection(locked=False)
    if number_format:
        cell.number_format = number_format


def style_formula(cell, cross_sheet: bool = False, number_format: str | None = None) -> None:
    cell.font = Font(name="Aptos", size=9, color=LINK_GREEN if cross_sheet else BLACK)
    if number_format:
        cell.number_format = number_format


def add_note(cell, text: str) -> None:
    cell.comment = Comment(text, "MWG-DMX portfolio model")


def setup_sheet(ws, freeze: str = "A4", zoom: int = 90) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.freeze_panes = freeze
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "MWG / DMX IPO SOTP — portfolio case study — not investment advice"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.sheet_properties.tabColor = TEAL


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_status_conditional_formatting(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'LEFT({cell_range.split(":")[0]},4)="PASS"'], fill=fill(PALE_GREEN)),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'LEFT({cell_range.split(":")[0]},4)="FAIL"'], fill=fill(PALE_RED)),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'LEFT({cell_range.split(":")[0]},4)="WARN"'], fill=fill(PALE_YELLOW)),
    )


def verify_embedded_q1_against_raw() -> list[str]:
    """Verify a representative set of embedded facts if the source pack exists."""
    if not RAW_Q1_PATH.exists():
        return ["Raw Q1 data pack absent (expected for a public clone); embedded sourced facts used."]

    wb = load_workbook(RAW_Q1_PATH, read_only=False, data_only=True)
    checks = {
        ("Income Statement", "Net revenue", 3): Q1["revenue_2026"],
        ("Income Statement", "Gross profit", 3): Q1["gross_profit_2026"],
        ("Income Statement", "Net profit after tax", 3): Q1["npat_2026"],
        ("Balance Sheet", "TOTAL ASSETS", 3): Q1["total_assets_mar_2026"],
        ("Balance Sheet", "TOTAL LIABILITIES & EQUITY", 3): Q1["total_assets_mar_2026"],
        ("Cash Flow Statement", "Net cash from operating activities", 3): Q1["cfo_2026"],
        ("Cash Flow Statement", "Cash and cash equipvalents at end of period", 3): Q1["cash_close_2026"],
    }
    messages: list[str] = []
    for (sheet, label, value_col), expected_bn in checks.items():
        ws = wb[sheet]
        found = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 2).value == label:
                found = ws.cell(row, value_col).value
                break
        if found is None:
            raise ValueError(f"Cannot find {label!r} in {sheet!r}")
        actual_bn = bn(found)
        if abs(actual_bn - expected_bn) > 0.005:
            raise ValueError(f"Embedded fact mismatch: {sheet}/{label}: {actual_bn} vs {expected_bn}")
        messages.append(f"Verified {sheet}/{label}")
    return messages


def new_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name in SHEET_ORDER:
        wb.create_sheet(name)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcId = 191029
    wb.properties.title = "MWG After the DMX IPO — DCF and SOTP Model"
    wb.properties.subject = "Portfolio financial model: IPO bridge, DMX DCF and MWG SOTP"
    wb.properties.creator = "Independent analyst work sample"
    wb.properties.keywords = "MWG, DMX, IPO, DCF, SOTP, Vietnam, financial model"
    wb.properties.description = (
        "Official facts, company outlook and illustrative assumptions are separated. "
        "No investment recommendation or market conclusion."
    )
    return wb


def build_cover(wb: Workbook) -> None:
    ws = wb["Cover"]
    setup_sheet(ws, "A1", 95)
    ws.page_setup.fitToHeight = 1
    ws.print_area = "B2:J41"
    ws.sheet_properties.tabColor = NAVY
    set_widths(ws, {"A": 4, "B": 23, "C": 20, "D": 18, "E": 4, "F": 4, "G": 24, "H": 24, "I": 22, "J": 22})

    ws.merge_cells("B2:J3")
    ws["B2"] = "MWG AFTER THE DMX IPO"
    ws["B2"].font = Font(name="Aptos Display", size=27, bold=True, color=WHITE)
    ws["B2"].fill = fill(NAVY)
    ws["B2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 28

    ws.merge_cells("B4:J5")
    ws["B4"] = "IPO bridge · DMX FCFF DCF · MWG sum-of-the-parts · scenario & QA model"
    ws["B4"].font = Font(name="Aptos", size=13, color="D9EAF7", italic=True)
    ws["B4"].fill = fill(NAVY_2)
    ws["B4"].alignment = Alignment(vertical="center")

    for r in range(2, 6):
        ws.cell(r, 1).fill = fill(NAVY)
    for r in range(2, 6):
        ws.cell(r, 10).fill = fill(NAVY if r <= 3 else NAVY_2)

    section(ws, 8, "MODEL CONTROL & PUBLICATION STATUS", 2, 10)
    controls = [
        ("Model version", MODEL_VERSION, "Formula and workbook architecture version"),
        ("Data cut-off", DATA_CUTOFF, "Official facts known at this date"),
        ("Generated at", GENERATED_AT, "Local build timestamp"),
        ("Active scenario", "=Assumptions!$C$4", "Bear / Base / Bull selector"),
        ("Display unit", "VND bn unless stated", "Share count in million; price in VND/share"),
        ("Valuation date", "N/A", "No same-date market price supplied"),
        ("Market conclusion", "UNAVAILABLE", "Guarded until price, price date and diluted shares are verified"),
    ]
    start = 9
    for i, (label, value, note) in enumerate(controls, start):
        ws.cell(i, 2, label).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        ws.cell(i, 3, value)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
        ws.cell(i, 3).fill = fill(LIGHT_BLUE if label != "Market conclusion" else PALE_YELLOW)
        ws.cell(i, 3).font = Font(
            name="Aptos",
            size=9,
            bold=label in {"Active scenario", "Market conclusion"},
            color=LINK_GREEN if isinstance(value, str) and value.startswith("=") else BLACK,
        )
        if isinstance(value, date):
            ws.cell(i, 3).number_format = FMT_DATE
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=10)
        ws.cell(i, 7, note).font = Font(name="Aptos", size=9, color=GREY)
        ws.cell(i, 7).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[i].height = 22

    section(ws, 18, "WHAT THIS MODEL ANSWERS", 2, 10)
    questions = [
        "1. How did the actual IPO change DMX shares and primary capital raised?",
        "2. What does an illustrative FCFF DCF imply for 100% of DMX under transparent drivers?",
        "3. How much of DMX's value is attributable to MWG using an explicitly labelled ownership proxy?",
        "4. What illustrative residual value is assigned to BHX and other MWG assets in a SOTP?",
        "5. Which assumptions drive the range, and which QA gates prevent an unsupported conclusion?",
    ]
    for row, question in enumerate(questions, 19):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.cell(row, 2, question).font = Font(name="Aptos", size=10, color=BLACK)
        ws.cell(row, 2).fill = fill(WHITE if row % 2 else LIGHT_GREY)
        ws.cell(row, 2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 24

    section(ws, 26, "WORKBOOK MAP", 2, 10)
    headers = ["Tab", "Purpose", "Primary status", "Reviewer focus"]
    for col, header in zip([2, 3, 7, 9], headers):
        ws.cell(27, col, header)
    ws.merge_cells("C27:F27")
    ws.merge_cells("G27:H27")
    ws.merge_cells("I27:J27")
    style_table_header(ws, 27, 2, 10)
    maps = [
        ("Sources", "Official source registry and URLs", "Official metadata", "Provenance"),
        ("IPO_Bridge", "Planned versus actual primary issuance", "Actual + planned", "Shares/proceeds identity"),
        ("Q1_Actuals", "Key consolidated statements and ratios", "Actual", "Statement/cash checks"),
        ("Mgmt_Forecast", "2025 actual and 2026–2030 company outlook", "Actual / company outlook", "No relabelling as analyst forecast"),
        ("Assumptions", "Editable Bear/Base/Bull inputs", "Illustrative", "Replace before investment use"),
        ("DMX_Valuation", "Driver-based FCFF DCF and equity bridge", "Illustrative forecast", "WACC, TV, double count"),
        ("MWG_SOTP", "Equity-level SOTP and holdco discount", "Illustrative", "Ownership and perimeter"),
        ("Scenarios", "Independent scenario calculations and sensitivities", "Illustrative", "Range, not target price"),
        ("Checks", "Critical QA and publication gates", "Control", "PASS/FAIL/WARN"),
    ]
    for r, row in enumerate(maps, 28):
        ws.cell(r, 2, row[0]).hyperlink = f"#'{row[0]}'!A1"
        ws.cell(r, 2).style = "Hyperlink"
        ws.cell(r, 3, row[1])
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.cell(r, 7, row[2])
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        ws.cell(r, 9, row[3])
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
        for c in range(2, 11):
            ws.cell(r, c).fill = fill(WHITE if r % 2 else LIGHT_GREY)
            ws.cell(r, c).border = Border(bottom=THIN_GREY)
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 28

    ws.merge_cells("B39:J41")
    ws["B39"] = (
        "DISCLAIMER — Portfolio case study only. Not investment research, a recommendation, or an offer to buy/sell securities. "
        "Official facts remain subject to source documents. Management outlook is not a guarantee. All yellow/blue inputs are "
        "illustrative analyst assumptions. No market upside/downside is shown because same-date price inputs are absent."
    )
    ws["B39"].fill = fill(PALE_YELLOW)
    ws["B39"].font = Font(name="Aptos", size=10, bold=True, color=NAVY)
    ws["B39"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["B39"].border = Border(top=MEDIUM_NAVY, bottom=MEDIUM_NAVY, left=MEDIUM_NAVY, right=MEDIUM_NAVY)


def build_sources(wb: Workbook) -> None:
    ws = wb["Sources"]
    setup_sheet(ws, "A5", 80)
    set_title(ws, "Source registry", "Official issuer documents used in the model; retrieved 13-Jul-2026", 13)
    set_widths(
        ws,
        {
            "A": 20,
            "B": 10,
            "C": 23,
            "D": 43,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 15,
            "I": 21,
            "J": 55,
            "K": 39,
            "L": 37,
            "M": 48,
        },
    )
    headers = [
        "Source ID",
        "Issuer",
        "Document type",
        "Title",
        "Period end",
        "Published",
        "Retrieved",
        "Audit status",
        "Scope",
        "Official URL",
        "Local artifact",
        "Model use",
        "Notes / limitations",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 1, len(headers))

    for row, src in enumerate(SOURCES, 5):
        values = [
            src.source_id,
            src.issuer,
            src.document_type,
            src.title,
            src.period_end,
            src.published_date,
            src.retrieved_at,
            src.audit_status,
            src.scope,
            src.url,
            src.local_artifact,
            src.use,
            src.notes,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        for col in (5, 6, 7):
            ws.cell(row, col).number_format = FMT_DATE
        ws.cell(row, 10).hyperlink = src.url
        ws.cell(row, 10).style = "Hyperlink"
        ws.row_dimensions[row].height = 54

    note_row = 5 + len(SOURCES) + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=13)
    ws.cell(note_row, 1, (
        "Repository policy: raw issuer documents are git-ignored. The build script embeds only the derived official facts needed "
        "for reproducibility and links each fact to the issuer source. If the local Q1 data pack is present, the build verifies "
        "representative embedded values against it before saving the model."
    ))
    ws.cell(note_row, 1).fill = fill(LIGHT_BLUE)
    ws.cell(note_row, 1).font = Font(name="Aptos", size=9, italic=True, color=NAVY)
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="center")


def build_ipo_bridge(wb: Workbook) -> dict[str, int]:
    ws = wb["IPO_Bridge"]
    setup_sheet(ws, "A5", 90)
    set_title(
        ws,
        "DMX IPO bridge — planned vs actual close",
        "Primary issuance; MWG did not receive the gross proceeds. Shares in million, values in VND bn unless stated.",
        8,
    )
    set_widths(ws, {"A": 34, "B": 15, "C": 18, "D": 18, "E": 20, "F": 25, "G": 52, "H": 24})

    headers = ["Metric", "Unit", "Planned", "Actual close", "Status", "Source / formula", "Interpretation", "Control"]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 1, 8)

    # Values in million shares make the unit conversion explicit: mn × VND / 1,000 = VND bn.
    rows: list[tuple[str, str, Any, Any, str, str, str]] = [
        ("IPO offer price", "VND/share", 80_000, 80_000, "Actual / planned", "DMX-IPO-SALESKIT; DMX-IPO-RES14", "Primary offer price"),
        ("Shares before IPO", "mn shares", 1_101.2835, 1_101.2835, "Actual", "DMX-IPO-RES14", "Opening share count"),
        ("Shares offered", "mn shares", 179.5004, 179.5004, "Actual / planned", "DMX-IPO-SALESKIT; DMX-IPO-RES14", "Maximum primary shares offered"),
        ("Shares registered", "mn shares", "N/A", 166.4858, "Actual", "DMX-IPO-RES14", "Registered, not necessarily allocated"),
        ("Shares successfully allocated", "mn shares", 179.5004, 166.4385, "Actual / planned", "DMX-IPO-RES14", "Actual case uses only paid/allocated shares"),
        ("Shares cancelled", "mn shares", "=C7-C9", 13.0619, "Actual", "DMX-IPO-RES14", "Unallocated shares were cancelled"),
        ("Shares after IPO", "mn shares", "=C6+C9", 1_267.7220, "Actual", "DMX-IPO-RES14", "Post-IPO diluted share count"),
        ("Gross proceeds", "VND bn", "=C9*C5/1000", "=D9*D5/1000", "Calculated from actual", "Shares allocated × offer price", "Cash raised by DMX before issue costs"),
        ("Subscription rate", "%", "=C9/C7", "=D9/D7", "Calculated", "Allocated / offered", "Actual take-up of the offer"),
        ("New shares / post-money shares", "%", "=C9/C11", "=D9/D11", "Calculated", "New / post shares", "Primary dilution percentage"),
        ("Pre-money equity value at IPO price", "VND bn", "=C6*C5/1000", "=D6*D5/1000", "Calculated", "Pre shares × offer price", "Transaction-price identity, not independent fair value"),
        ("Post-money equity value at IPO price", "VND bn", "=C11*C5/1000", "=D11*D5/1000", "Calculated", "Post shares × offer price", "Transaction-price identity, not a market conclusion"),
        ("MWG ownership after IPO", "%", "N/A", 0.86, "Approximate disclosure", "DMX-IPO-COMPLETE", "'Nearly 86%' only; exact shares remain unverified"),
        ("Issue costs", "VND bn", "N/A", "N/A", "Unavailable", "Not disclosed precisely in current source set", "Do not guess"),
        ("Net proceeds", "VND bn", "N/A", '=IF(ISNUMBER(D18),D12-D18,"N/A")', "Unavailable until issue costs verified", "Gross proceeds − issue costs", "No double count in equity bridge"),
        ("Additional IPO proceeds in DCF bridge", "VND bn", 0.0, 0.0, "Control", "Post-IPO balance sheet convention", "Must remain zero because proceeds enter cash once"),
    ]
    row_map: dict[str, int] = {}
    for row, item in enumerate(rows, 5):
        label, unit, planned, actual, status, source, interpretation = item
        row_map[label] = row
        values = [label, unit, planned, actual, status, source, interpretation]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        for col in (3, 4):
            value = ws.cell(row, col).value
            if unit == "%":
                ws.cell(row, col).number_format = FMT_PCT_2
            elif unit == "VND/share":
                ws.cell(row, col).number_format = FMT_VND
            elif unit == "mn shares":
                ws.cell(row, col).number_format = FMT_SHARES
            else:
                ws.cell(row, col).number_format = FMT_BN_2
            if isinstance(value, str) and value.startswith("="):
                style_formula(ws.cell(row, col), False, ws.cell(row, col).number_format)
            elif isinstance(value, (int, float)):
                ws.cell(row, col).font = Font(name="Aptos", size=9, color=INPUT_BLUE)
        ws.row_dimensions[row].height = 30

    # Formula-based controls beside the bridge.
    control_formulas = {
        5: '=IF(C5=D5,"PASS","FAIL")',
        10: '=IF(ABS(D7-D9-D10)<0.000001,"PASS","FAIL")',
        11: '=IF(ABS((D6+D9)-D11)<0.000001,"PASS","FAIL")',
        12: '=IF(ABS(D9*D5/1000-13315.080)<0.001,"PASS","FAIL")',
        16: '=IF(ABS(D15+D12-D16)<0.001,"PASS","FAIL")',
        17: '=IF(ABS(D17-86%)<=1%,"PASS — rounded","WARN")',
        20: '=IF(D20=0,"PASS","FAIL")',
    }
    for row, formula in control_formulas.items():
        ws.cell(row, 8, formula)
        ws.cell(row, 8).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 8).font = Font(name="Aptos", size=9, bold=True)
    add_status_conditional_formatting(ws, "H5:H20")

    note_row = 23
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=8)
    ws.cell(note_row, 1, (
        "Accounting boundary: this is a primary issuance. Gross proceeds belong to DMX, not MWG. Under the post-IPO balance-sheet "
        "convention used in the DCF, cash/debt are taken from the latest available actual balance sheet and additional IPO proceeds "
        "remain zero. Exact MWG share ownership is not inferred from the rounded 'nearly 86%' announcement."
    ))
    ws.cell(note_row, 1).fill = fill(PALE_YELLOW)
    ws.cell(note_row, 1).font = Font(name="Aptos", size=9, color=NAVY, bold=True)
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    return row_map


def build_q1_actuals(wb: Workbook) -> dict[str, int]:
    ws = wb["Q1_Actuals"]
    setup_sheet(ws, "A5", 85)
    set_title(
        ws,
        "DMX Q1 2026 actuals",
        "Consolidated issuer data pack; reported signs preserved. P&L/CF compare Q1; BS compares 31-Mar-2026 with 31-Dec-2025.",
        10,
    )
    set_widths(ws, {"A": 12, "B": 36, "C": 13, "D": 18, "E": 18, "F": 14, "G": 19, "H": 28, "I": 16, "J": 45})
    headers = ["Statement", "Line item", "Unit", "Current period", "Comparator", "Change / ratio", "Source ID", "Source locator", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 1, 10)

    row_map: dict[str, int] = {}
    row = 5

    def add_section(label: str, period_note: str) -> None:
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.cell(row, 1, f"{label} — {period_note}")
        ws.cell(row, 1).fill = fill(NAVY_2)
        ws.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        ws.cell(row, 1).alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

    def add_line(
        key: str,
        statement: str,
        label: str,
        unit: str,
        current: Any,
        comparator: Any,
        variance: str | None,
        source_id: str,
        locator: str,
        status: str,
        notes: str = "",
    ) -> None:
        nonlocal row
        row_map[key] = row
        values = [statement, label, unit, current, comparator, variance, source_id, locator, status, notes]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        fmt = FMT_VND if unit == "VND/share" else (FMT_PCT if unit == "%" else FMT_BN_1)
        for col in (4, 5, 6):
            ws.cell(row, col).number_format = fmt
            if isinstance(ws.cell(row, col).value, str) and ws.cell(row, col).value.startswith("="):
                style_formula(ws.cell(row, col), False, fmt)
        if status == "Actual":
            ws.cell(row, 4).font = Font(name="Aptos", size=9, color=INPUT_BLUE)
            ws.cell(row, 5).font = Font(name="Aptos", size=9, color=INPUT_BLUE)
        elif status == "Calculated":
            ws.cell(row, 1).fill = fill(LIGHT_TEAL)
            ws.cell(row, 2).fill = fill(LIGHT_TEAL)
        ws.row_dimensions[row].height = 27
        row += 1

    add_section("INCOME STATEMENT", "Q1 2026 vs Q1 2025 | VND bn except EPS and ratios")
    add_line("revenue", "IS", "Net revenue", "VND bn", Q1["revenue_2026"], Q1["revenue_2025"], "=ABS(D6)/ABS(E6)-1", "DMX-Q1-2026", "Income Statement — Net revenue", "Actual")
    add_line("cogs", "IS", "Cost of goods sold", "VND bn", Q1["cogs_2026"], Q1["cogs_2025"], "=ABS(D7)/ABS(E7)-1", "DMX-Q1-2026", "Income Statement — Cost of goods sold", "Actual", "Reported as a negative expense")
    add_line("gross_profit", "IS", "Gross profit", "VND bn", Q1["gross_profit_2026"], Q1["gross_profit_2025"], "=D8/E8-1", "DMX-Q1-2026", "Income Statement — Gross profit", "Actual")
    add_line("gross_margin", "IS", "Gross margin", "%", "=D8/D6", "=E8/E6", "=D9-E9", "MODEL-FORMULA", "Gross profit / net revenue", "Calculated", "Change shown in percentage points")
    add_line("financial_income", "IS", "Financial income", "VND bn", Q1["financial_income_2026"], Q1["financial_income_2025"], "=D10/E10-1", "DMX-Q1-2026", "Income Statement — Financial income", "Actual")
    add_line("financial_expense", "IS", "Financial expense", "VND bn", Q1["financial_expense_2026"], Q1["financial_expense_2025"], "=ABS(D11)/ABS(E11)-1", "DMX-Q1-2026", "Income Statement — Financial expense", "Actual", "Reported as a negative expense")
    add_line("interest_expense", "IS", "Interest expense", "VND bn", Q1["interest_expense_2026"], Q1["interest_expense_2025"], "=ABS(D12)/ABS(E12)-1", "DMX-Q1-2026", "Income Statement — Interest expense", "Actual")
    add_line("selling_expense", "IS", "Selling expense", "VND bn", Q1["selling_expense_2026"], Q1["selling_expense_2025"], "=ABS(D13)/ABS(E13)-1", "DMX-Q1-2026", "Income Statement — Selling expense", "Actual")
    add_line("ga_expense", "IS", "G&A expense", "VND bn", Q1["ga_expense_2026"], Q1["ga_expense_2025"], "=ABS(D14)/ABS(E14)-1", "DMX-Q1-2026", "Income Statement — G&A expense", "Actual")
    add_line("share_jv", "IS", "Share of profit from JV", "VND bn", Q1["share_jv_2026"], Q1["share_jv_2025"], "=D15/E15-1", "DMX-Q1-2026", "Income Statement — Share of JV profit", "Actual", "EraBlue is treated as a JV perimeter control")
    add_line("operating_profit", "IS", "Operating profit", "VND bn", Q1["operating_profit_2026"], Q1["operating_profit_2025"], "=D16/E16-1", "DMX-Q1-2026", "Income Statement — Operating profit", "Actual")
    add_line("operating_margin", "IS", "Operating margin", "%", "=D16/D6", "=E16/E6", "=D17-E17", "MODEL-FORMULA", "Operating profit / revenue", "Calculated", "Company line includes finance and JV effects; not used as DCF EBIT")
    add_line("pbt", "IS", "Profit before tax", "VND bn", Q1["pbt_2026"], Q1["pbt_2025"], "=D18/E18-1", "DMX-Q1-2026", "Income Statement — Profit before tax", "Actual")
    add_line("tax", "IS", "Income tax expense", "VND bn", Q1["tax_2026"], Q1["tax_2025"], "=ABS(D19)/ABS(E19)-1", "DMX-Q1-2026", "Income Statement — current + deferred tax", "Actual")
    add_line("npat", "IS", "Net profit after tax", "VND bn", Q1["npat_2026"], Q1["npat_2025"], "=D20/E20-1", "DMX-Q1-2026", "Income Statement — NPAT", "Actual")
    add_line("npat_margin", "IS", "NPAT margin", "%", "=D20/D6", "=E20/E6", "=D21-E21", "MODEL-FORMULA", "NPAT / revenue", "Calculated", "Change shown in percentage points")
    add_line("eps", "IS", "Basic EPS", "VND/share", Q1["eps_2026"], Q1["eps_2025"], "=D22/E22-1", "DMX-Q1-2026", "Income Statement — Basic EPS", "Actual")

    row += 1
    add_section("BALANCE SHEET", "31-Mar-2026 vs 31-Dec-2025 | VND bn")
    bs_items = [
        ("cash", "Cash & cash equivalents", Q1["cash_mar_2026"], Q1["cash_dec_2025"], "Balance Sheet — Cash & cash equivalents"),
        ("short_term_investments", "Short-term held-to-maturity investments", Q1["short_term_investments_mar_2026"], Q1["short_term_investments_dec_2025"], "Balance Sheet — Short-term investments"),
        ("long_term_deposits", "Long-term held-to-maturity investments", Q1["long_term_deposits_mar_2026"], Q1["long_term_deposits_dec_2025"], "Balance Sheet — Long-term held-to-maturity investments"),
        ("cash_deposits", "Cash + term deposits", None, None, "Cash + short-term investments + long-term deposits"),
        ("receivables", "Current receivables", Q1["receivables_mar_2026"], Q1["receivables_dec_2025"], "Balance Sheet — Current receivables"),
        ("inventory", "Inventory, net", Q1["inventory_mar_2026"], Q1["inventory_dec_2025"], "Balance Sheet — Inventories"),
        ("inventory_provision", "Inventory provision", Q1["inventory_provision_mar_2026"], Q1["inventory_provision_dec_2025"], "Balance Sheet — Inventory provision"),
        ("ppe", "Fixed assets, net", Q1["ppe_mar_2026"], Q1["ppe_dec_2025"], "Balance Sheet — Fixed assets"),
        ("jv_investment", "Investment in jointly controlled entity", Q1["jv_investment_mar_2026"], Q1["jv_investment_dec_2025"], "Balance Sheet — JV investment"),
        ("total_assets", "Total assets", Q1["total_assets_mar_2026"], Q1["total_assets_dec_2025"], "Balance Sheet — Total assets"),
        ("payables", "Short-term trade payables", Q1["payables_mar_2026"], Q1["payables_dec_2025"], "Balance Sheet — Trade payables"),
        ("debt", "Short-term debt", Q1["short_term_debt_mar_2026"], Q1["short_term_debt_dec_2025"], "Balance Sheet — Short-term loans"),
        ("total_liabilities", "Total liabilities", Q1["total_liabilities_mar_2026"], Q1["total_liabilities_dec_2025"], "Balance Sheet — Total liabilities"),
        ("equity", "Total equity", Q1["equity_mar_2026"], Q1["equity_dec_2025"], "Balance Sheet — Equity"),
        ("net_cash", "Cash + deposits − debt", None, None, "Calculated liquidity bridge"),
    ]
    for key, label, current, prior, locator in bs_items:
        current_value: Any = current
        prior_value: Any = prior
        status = "Actual"
        source_id = "DMX-Q1-2026"
        if key == "cash_deposits":
            current_value = f'=D{row_map["cash"]}+D{row_map["short_term_investments"]}+D{row_map["long_term_deposits"]}'
            prior_value = f'=E{row_map["cash"]}+E{row_map["short_term_investments"]}+E{row_map["long_term_deposits"]}'
            status = "Calculated"
            source_id = "MODEL-FORMULA"
        elif key == "net_cash":
            current_value = f'=D{row_map["cash_deposits"]}-D{row_map["debt"]}'
            prior_value = f'=E{row_map["cash_deposits"]}-E{row_map["debt"]}'
            status = "Calculated"
            source_id = "MODEL-FORMULA"
        add_line(key, "BS", label, "VND bn", current_value, prior_value, f'=IFERROR(D{row}/E{row}-1,"N/A")', source_id, locator, status)

    row += 1
    add_section("CASH FLOW STATEMENT", "Q1 2026 vs Q1 2025 | VND bn")
    cf_items = [
        ("cfo", "Net cash from operating activities", Q1["cfo_2026"], Q1["cfo_2025"], "Cash Flow — CFO"),
        ("capex", "Purchase/construction of fixed assets", Q1["capex_2026"], Q1["capex_2025"], "Cash Flow — Capex"),
        ("fcf_actual", "FCF after capex", None, None, "CFO + capex (capex reported negative)"),
        ("cfi", "Net cash from investing activities", Q1["cfi_2026"], Q1["cfi_2025"], "Cash Flow — CFI"),
        ("debt_drawdown", "Drawdown from borrowings", Q1["debt_drawdown_2026"], Q1["debt_drawdown_2025"], "Cash Flow — debt drawdown"),
        ("debt_repayment", "Repayment of borrowings", Q1["debt_repayment_2026"], Q1["debt_repayment_2025"], "Cash Flow — debt repayment"),
        ("cff", "Net cash from financing activities", Q1["cff_2026"], Q1["cff_2025"], "Cash Flow — CFF"),
        ("net_cash_change", "Net change in cash", Q1["net_cash_change_2026"], Q1["net_cash_change_2025"], "Cash Flow — net change in cash"),
        ("cash_open", "Opening cash", Q1["cash_open_2026"], Q1["cash_open_2025"], "Cash Flow — opening cash"),
        ("fx_effect", "FX effect", Q1["fx_effect_2026"], Q1["fx_effect_2025"], "Cash Flow — FX effect"),
        ("cash_close", "Closing cash", Q1["cash_close_2026"], Q1["cash_close_2025"], "Cash Flow — closing cash"),
        ("cfo_conversion", "CFO / NPAT", None, None, "CFO / NPAT"),
    ]
    for key, label, current, prior, locator in cf_items:
        status = "Actual"
        source_id = "DMX-Q1-2026"
        unit = "VND bn"
        current_value: Any = current
        prior_value: Any = prior
        if key == "fcf_actual":
            current_value = f'=D{row_map["cfo"]}+D{row_map["capex"]}'
            prior_value = f'=E{row_map["cfo"]}+E{row_map["capex"]}'
            status = "Calculated"
            source_id = "MODEL-FORMULA"
        elif key == "cfo_conversion":
            current_value = f'=D{row_map["cfo"]}/D{row_map["npat"]}'
            prior_value = f'=E{row_map["cfo"]}/E{row_map["npat"]}'
            status = "Calculated"
            source_id = "MODEL-FORMULA"
            unit = "%"
        add_line(key, "CF", label, unit, current_value, prior_value, f'=IFERROR(D{row}/E{row}-1,"N/A")', source_id, locator, status)

    # Small presentation chart built from a hidden helper table.
    helper_col = 12
    ws.cell(2, helper_col, "Metric")
    ws.cell(2, helper_col + 1, "Q1 2026")
    ws.cell(2, helper_col + 2, "Q1 2025")
    for idx, (label, key) in enumerate([("Revenue", "revenue"), ("Gross profit", "gross_profit"), ("NPAT", "npat")], 3):
        ws.cell(idx, helper_col, label)
        ws.cell(idx, helper_col + 1, f'=D{row_map[key]}')
        ws.cell(idx, helper_col + 2, f'=E{row_map[key]}')
    for col in range(helper_col, helper_col + 3):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Q1 operating snapshot"
    chart.y_axis.title = "VND bn"
    chart.height = 7.2
    chart.width = 13.5
    data = Reference(ws, min_col=helper_col + 1, max_col=helper_col + 2, min_row=2, max_row=5)
    cats = Reference(ws, min_col=helper_col, min_row=3, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "L7")

    ws.auto_filter.ref = f"A4:J{row - 1}"
    return row_map


def build_mgmt_forecast(wb: Workbook) -> dict[str, int]:
    ws = wb["Mgmt_Forecast"]
    setup_sheet(ws, "C5", 90)
    set_title(
        ws,
        "DMX management outlook",
        "2025 actual shown in company material; 2026F–2030F are management outlook, not analyst estimates and not guarantees.",
        10,
    )
    set_widths(ws, {"A": 4, "B": 34, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 25, "J": 45})

    years = list(MGMT_FORECAST)
    ws.cell(4, 2, "Metric")
    for col, year in enumerate(years, 3):
        ws.cell(4, col, f"{year}{'A' if year == 2025 else 'F'}")
    ws.cell(4, 9, "Source / status")
    ws.cell(4, 10, "Interpretation")
    style_table_header(ws, 4, 2, 10)

    row_map = {
        "status": 5,
        "revenue": 6,
        "revenue_growth": 7,
        "npat": 8,
        "npat_growth": 9,
        "npat_margin": 10,
        "gross_margin": 11,
    }
    labels = {
        "status": "Data status",
        "revenue": "Revenue",
        "revenue_growth": "Revenue growth",
        "npat": "NPAT",
        "npat_growth": "NPAT growth",
        "npat_margin": "NPAT margin",
        "gross_margin": "Gross margin",
    }
    for key, row in row_map.items():
        ws.cell(row, 2, labels[key])
        for col, year in enumerate(years, 3):
            if key == "status":
                value: Any = MGMT_FORECAST[year]["status"]
            elif key == "revenue":
                value = MGMT_FORECAST[year]["revenue"]
            elif key == "npat":
                value = MGMT_FORECAST[year]["npat"]
            elif key == "gross_margin":
                value = MGMT_FORECAST[year]["gross_margin"]
            elif key == "revenue_growth":
                value = "N/A" if year == 2025 else f"={get_column_letter(col)}6/{get_column_letter(col - 1)}6-1"
            elif key == "npat_growth":
                value = "N/A" if year == 2025 else f"={get_column_letter(col)}8/{get_column_letter(col - 1)}8-1"
            else:
                value = f"={get_column_letter(col)}8/{get_column_letter(col)}6"
            ws.cell(row, col, value)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if key in {"revenue", "npat", "gross_margin"}:
                ws.cell(row, col).font = Font(name="Aptos", size=9, color=INPUT_BLUE)
            else:
                style_formula(ws.cell(row, col), False)
            if key in {"revenue_growth", "npat_growth", "npat_margin", "gross_margin"}:
                ws.cell(row, col).number_format = FMT_PCT
            elif key in {"revenue", "npat"}:
                ws.cell(row, col).number_format = FMT_BN
        ws.cell(row, 9, "DMX-OUTLOOK")
        ws.cell(row, 10, "Company outlook is held separate from the analyst scenario layer.")
        for col in (2, 9, 10):
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).font = Font(name="Aptos", size=9, bold=col == 2, color=NAVY if col == 2 else BLACK)
        ws.row_dimensions[row].height = 26

    ws.merge_cells("B14:J16")
    ws["B14"] = (
        "MODEL BOUNDARY — These values reproduce the company's published outlook as captured at the data cut-off. The DCF does not "
        "silently treat them as audited actuals: it applies a separate scenario growth adjustment and explicit EBIT, tax, D&A, capex "
        "and working-capital assumptions on the Assumptions tab."
    )
    ws["B14"].fill = fill(PALE_YELLOW)
    ws["B14"].font = Font(name="Aptos", size=9, bold=True, color=NAVY)
    ws["B14"].alignment = Alignment(wrap_text=True, vertical="center")

    chart = LineChart()
    chart.title = "Company outlook: revenue and NPAT"
    chart.y_axis.title = "VND bn"
    chart.x_axis.title = "Year"
    chart.style = 13
    chart.height = 8.2
    chart.width = 15.5
    for source_row in (6, 8):
        data = Reference(ws, min_col=2, max_col=8, min_row=source_row, max_row=source_row)
        chart.add_data(data, titles_from_data=True, from_rows=True)
    categories = Reference(ws, min_col=3, max_col=8, min_row=4, max_row=4)
    chart.set_categories(categories)
    ws.add_chart(chart, "B19")
    return row_map


def build_assumptions(wb: Workbook) -> dict[str, int]:
    ws = wb["Assumptions"]
    setup_sheet(ws, "A11", 85)
    set_title(
        ws,
        "Illustrative analyst assumptions",
        "Yellow cells / blue font are editable inputs. None of the scenario values below is an official company forecast unless separately sourced.",
        11,
    )
    set_widths(ws, {"A": 23, "B": 21, "C": 37, "D": 14, "E": 15, "F": 15, "G": 15, "H": 16, "I": 22, "J": 52, "K": 27})

    ws["A4"] = "Active scenario"
    ws["C4"] = "Base"
    style_input(ws["C4"])
    ws["C4"].font = Font(name="Aptos", size=11, bold=True, color=INPUT_BLUE)
    dv = DataValidation(type="list", formula1='"Bear,Base,Bull"', allow_blank=False)
    dv.error = "Select Bear, Base or Bull."
    dv.errorTitle = "Invalid scenario"
    dv.prompt = "Select the active operating and valuation scenario."
    dv.promptTitle = "Scenario"
    ws.add_data_validation(dv)
    dv.add(ws["C4"])

    controls = [
        (5, "Valuation date", "N/A", "Required for a market conclusion"),
        (6, "MWG closing price", "N/A", "VND/share; must have the same valuation date"),
        (7, "DMX market price", "N/A", "N/A before verified trading price/date"),
        (8, "Exact MWG ownership of DMX", "N/A", "Rounded 'nearly 86%' is not exact ownership"),
    ]
    for r, label, value, note in controls:
        ws.cell(r, 1, label).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        ws.cell(r, 3, value).fill = fill(PALE_YELLOW)
        ws.cell(r, 3).font = Font(name="Aptos", size=9, color=INPUT_BLUE)
        ws.cell(r, 10, note).font = Font(name="Aptos", size=9, color=GREY)

    headers = ["Assumption ID", "Category", "Driver", "Unit", "Bear", "Base", "Bull", "Active", "Status", "Rationale", "Evidence / source"]
    for col, header in enumerate(headers, 1):
        ws.cell(10, col, header)
    style_table_header(ws, 10, 1, 11)

    assumptions: list[tuple[str, str, str, str, float, float, float, str, str, str]] = [
        ("OP_REV_GROWTH_ADJ", "DMX operations", "Annual adjustment to management revenue growth", "%", -0.030, 0.000, 0.020, "Illustrative", "Applied to the growth rate, not directly to revenue output", "Analyst scenario"),
        ("OP_EBIT_MARGIN_2026", "DMX operations", "EBIT margin in 2026E", "%", 0.070, 0.080, 0.088, "Illustrative", "FCFF EBIT is independent from company NPAT outlook", "Analyst scenario"),
        ("OP_EBIT_MARGIN_STEP", "DMX operations", "Annual EBIT margin step", "%", 0.0010, 0.0020, 0.0025, "Illustrative", "Linear step from the 2026E margin", "Analyst scenario"),
        ("OP_CASH_TAX", "DMX operations", "Cash tax rate", "%", 0.200, 0.200, 0.200, "Illustrative", "Vietnam corporate-tax simplification; replace with modelled cash tax", "Analyst scenario"),
        ("OP_DA_PCT_REV", "DMX operations", "D&A as % of revenue", "%", 0.010, 0.010, 0.010, "Illustrative", "Required to bridge EBIT to FCFF", "Analyst scenario"),
        ("OP_CAPEX_PCT_REV", "DMX operations", "Capex as % of revenue", "%", 0.012, 0.010, 0.009, "Illustrative", "Maintenance and growth capex combined", "Analyst scenario"),
        ("OP_NWC_PCT_DELTA_REV", "DMX operations", "Cash investment in NWC as % of revenue increase", "%", 0.040, 0.030, 0.020, "Illustrative", "Positive value is a cash outflow", "Analyst scenario"),
        ("VAL_WACC", "DMX valuation", "Nominal VND WACC", "%", 0.130, 0.120, 0.110, "Illustrative", "Market inputs are not populated; sensitivity is mandatory", "Analyst scenario"),
        ("VAL_TERMINAL_G", "DMX valuation", "Terminal growth", "%", 0.030, 0.040, 0.045, "Illustrative", "Must remain below WACC", "Analyst scenario"),
        ("VAL_ERABLUE", "DMX valuation", "Fair value of DMX stake in EraBlue", "VND bn", 1_000.0, 1_500.0, 2_000.0, "Illustrative", "Added once because FCFF EBIT excludes JV earnings", "MWG-ERABLUE-JV / analyst scenario"),
        ("VAL_LEASE", "DMX valuation", "Lease liabilities in equity bridge", "VND bn", 0.0, 0.0, 0.0, "Illustrative placeholder", "Replace after verifying the chosen lease convention", "Analyst placeholder"),
        ("VAL_OTHER_CLAIMS", "DMX valuation", "Other claims in equity bridge", "VND bn", 0.0, 0.0, 0.0, "Illustrative placeholder", "No unsupported claim is inserted", "Analyst placeholder"),
        ("SOTP_DMX_OWNERSHIP", "MWG SOTP", "MWG effective ownership of DMX", "%", 0.860, 0.860, 0.860, "Illustrative proxy", "Rounded disclosure only; exact shares required for production", "DMX-IPO-COMPLETE"),
        ("SOTP_BHX_REVENUE", "MWG SOTP", "BHX reference revenue", "VND bn", 43_000.0, 46_000.0, 49_000.0, "Illustrative", "Portfolio model placeholder", "Analyst scenario"),
        ("SOTP_BHX_EVS", "MWG SOTP", "BHX EV / Sales", "x", 0.60, 0.80, 1.00, "Illustrative", "Comparable analysis not yet populated", "Analyst scenario"),
        ("SOTP_BHX_NET_DEBT", "MWG SOTP", "BHX net debt", "VND bn", 3_000.0, 2_000.0, 1_000.0, "Illustrative", "Subsidiary-level bridge only", "Analyst scenario"),
        ("SOTP_BHX_OWNERSHIP", "MWG SOTP", "MWG ownership of BHX", "%", 0.950, 0.950, 0.950, "Illustrative", "Replace with exact effective ownership", "Analyst scenario"),
        ("SOTP_AK_REVENUE", "MWG SOTP", "An Khang reference revenue", "VND bn", 1_800.0, 2_200.0, 2_600.0, "Illustrative", "Portfolio model placeholder", "Analyst scenario"),
        ("SOTP_AK_EVS", "MWG SOTP", "An Khang EV / Sales", "x", 0.25, 0.35, 0.45, "Illustrative", "Loss-making / early-stage range", "Analyst scenario"),
        ("SOTP_AK_NET_DEBT", "MWG SOTP", "An Khang net debt", "VND bn", 500.0, 500.0, 500.0, "Illustrative", "Subsidiary-level bridge only", "Analyst scenario"),
        ("SOTP_AK_OWNERSHIP", "MWG SOTP", "MWG ownership of An Khang", "%", 1.000, 1.000, 1.000, "Illustrative", "Replace with exact effective ownership", "Analyst scenario"),
        ("SOTP_AVA_REVENUE", "MWG SOTP", "AVAKids reference revenue", "VND bn", 1_000.0, 1_200.0, 1_400.0, "Illustrative", "Portfolio model placeholder", "Analyst scenario"),
        ("SOTP_AVA_EVS", "MWG SOTP", "AVAKids EV / Sales", "x", 0.40, 0.60, 0.80, "Illustrative", "Early-stage range", "Analyst scenario"),
        ("SOTP_AVA_NET_DEBT", "MWG SOTP", "AVAKids net debt", "VND bn", 200.0, 200.0, 200.0, "Illustrative", "Subsidiary-level bridge only", "Analyst scenario"),
        ("SOTP_AVA_OWNERSHIP", "MWG SOTP", "MWG ownership of AVAKids", "%", 1.000, 1.000, 1.000, "Illustrative", "Replace with exact effective ownership", "Analyst scenario"),
        ("SOTP_OTHER_INV", "MWG SOTP", "Other investments", "VND bn", 500.0, 1_000.0, 1_500.0, "Illustrative", "Equity value, no separate net-debt adjustment", "Analyst scenario"),
        ("SOTP_PARENT_ASSETS", "MWG SOTP", "Parent non-operating assets", "VND bn", 0.0, 0.0, 0.0, "Illustrative placeholder", "Avoids unsupported double count", "Analyst placeholder"),
        ("SOTP_PARENT_NET_DEBT", "MWG SOTP", "Parent-only net debt", "VND bn", 2_500.0, 2_000.0, 1_500.0, "Illustrative", "Must not use consolidated debt after valuing subs at equity", "Analyst scenario"),
        ("SOTP_PARENT_PROVISIONS", "MWG SOTP", "Parent provisions / other claims", "VND bn", 1_000.0, 500.0, 250.0, "Illustrative", "Separate parent claims", "Analyst scenario"),
        ("SOTP_HOLDCO_DISC", "MWG SOTP", "Holding-company discount", "%", 0.250, 0.150, 0.050, "Illustrative", "Sensitivity, not a plug to a target price", "Analyst scenario"),
        ("SOTP_MWG_SHARES", "MWG SOTP", "MWG diluted shares", "mn shares", 1_477.2, 1_477.2, 1_477.2, "Illustrative placeholder", "Replace with verified diluted shares at valuation date", "Analyst placeholder"),
    ]

    row_map: dict[str, int] = {}
    for row, item in enumerate(assumptions, 11):
        assumption_id, category, driver, unit, bear, base, bull, status, rationale, evidence = item
        row_map[assumption_id] = row
        values = [assumption_id, category, driver, unit, bear, base, bull]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        for col in (5, 6, 7):
            style_input(ws.cell(row, col))
        ws.cell(row, 8, f'=INDEX(E{row}:G{row},1,MATCH($C$4,$E$10:$G$10,0))')
        ws.cell(row, 9, status)
        ws.cell(row, 10, rationale)
        ws.cell(row, 11, evidence)
        for col in range(8, 12):
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).fill = fill(LIGHT_TEAL if col == 8 else (WHITE if row % 2 else LIGHT_GREY))
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        if unit == "%":
            fmt = FMT_PCT
        elif unit == "x":
            fmt = FMT_MULTIPLE
        elif unit == "mn shares":
            fmt = FMT_SHARES
        else:
            fmt = FMT_BN
        for col in (5, 6, 7, 8):
            ws.cell(row, col).number_format = fmt
        add_note(ws.cell(row, 8), f"Active {driver}. Status: {status}. {rationale}")
        ws.row_dimensions[row].height = 32

    end = 10 + len(assumptions)
    ws.auto_filter.ref = f"A10:K{end}"
    return row_map


def build_dmx_valuation(
    wb: Workbook,
    q1_rows: dict[str, int],
    mgmt_rows: dict[str, int],
    assumption_rows: dict[str, int],
) -> dict[str, int]:
    ws = wb["DMX_Valuation"]
    setup_sheet(ws, "C5", 85)
    set_title(
        ws,
        "DMX illustrative FCFF DCF",
        "Management revenue path + explicit analyst EBIT/tax/D&A/capex/NWC drivers. Mid-year convention; post-IPO balance-sheet bridge.",
        13,
    )
    set_widths(ws, {"A": 4, "B": 38, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15, "I": 4, "J": 34, "K": 20, "L": 24, "M": 45})

    years = list(MGMT_FORECAST)
    ws.cell(4, 2, "FCFF build")
    for col, year in enumerate(years, 3):
        ws.cell(4, col, f"{year}{'A' if year == 2025 else 'E'}")
    style_table_header(ws, 4, 2, 8)

    rows = {
        "status": 5,
        "mgmt_revenue": 6,
        "mgmt_growth": 7,
        "model_revenue": 8,
        "growth_adjustment": 9,
        "mgmt_gross_margin": 10,
        "mgmt_npat": 11,
        "ebit_margin": 12,
        "ebit": 13,
        "cash_tax": 14,
        "nopat": 15,
        "da_pct": 16,
        "da": 17,
        "capex_pct": 18,
        "capex": 19,
        "delta_nwc": 20,
        "fcff": 21,
        "discount_period": 22,
        "discount_factor": 23,
        "pv_fcff": 24,
    }
    labels = {
        "status": "Data status",
        "mgmt_revenue": "Management revenue path",
        "mgmt_growth": "Management revenue growth",
        "model_revenue": "Model revenue",
        "growth_adjustment": "Scenario growth adjustment",
        "mgmt_gross_margin": "Management gross margin reference",
        "mgmt_npat": "Management NPAT reference",
        "ebit_margin": "Illustrative EBIT margin",
        "ebit": "EBIT",
        "cash_tax": "Cash tax rate",
        "nopat": "NOPAT",
        "da_pct": "D&A / revenue",
        "da": "D&A",
        "capex_pct": "Capex / revenue",
        "capex": "Capex (cash outflow)",
        "delta_nwc": "Change in NWC (cash outflow)",
        "fcff": "FCFF",
        "discount_period": "Discount period",
        "discount_factor": "Discount factor",
        "pv_fcff": "Present value of FCFF",
    }
    percent_rows = {"mgmt_growth", "growth_adjustment", "mgmt_gross_margin", "ebit_margin", "cash_tax", "da_pct", "capex_pct", "discount_factor"}
    for key, row in rows.items():
        ws.cell(row, 2, labels[key])
        ws.cell(row, 2).font = Font(name="Aptos", size=9, bold=key in {"model_revenue", "ebit", "fcff", "pv_fcff"}, color=NAVY)
        for col, year in enumerate(years, 3):
            letter = get_column_letter(col)
            mgmt_letter = get_column_letter(col)
            if key == "status":
                value: Any = "Actual / company presentation" if year == 2025 else "Illustrative forecast"
            elif key == "mgmt_revenue":
                value = f'=Mgmt_Forecast!{mgmt_letter}{mgmt_rows["revenue"]}'
            elif key == "mgmt_growth":
                value = f'=Mgmt_Forecast!{mgmt_letter}{mgmt_rows["revenue_growth"]}' if year > 2025 else "N/A"
            elif key == "growth_adjustment":
                value = f'=Assumptions!$H${assumption_rows["OP_REV_GROWTH_ADJ"]}' if year > 2025 else "N/A"
            elif key == "model_revenue":
                if year == 2025:
                    value = f'={letter}{rows["mgmt_revenue"]}'
                else:
                    prev = get_column_letter(col - 1)
                    value = f'={prev}{row}*(1+{letter}{rows["mgmt_growth"]}+{letter}{rows["growth_adjustment"]})'
            elif key == "mgmt_gross_margin":
                value = f'=Mgmt_Forecast!{mgmt_letter}{mgmt_rows["gross_margin"]}'
            elif key == "mgmt_npat":
                value = f'=Mgmt_Forecast!{mgmt_letter}{mgmt_rows["npat"]}'
            elif key == "ebit_margin":
                if year == 2025:
                    value = "N/A"
                elif year == 2026:
                    value = f'=Assumptions!$H${assumption_rows["OP_EBIT_MARGIN_2026"]}'
                else:
                    prev = get_column_letter(col - 1)
                    value = f'={prev}{row}+Assumptions!$H${assumption_rows["OP_EBIT_MARGIN_STEP"]}'
            elif key == "ebit":
                value = "N/A" if year == 2025 else f'={letter}{rows["model_revenue"]}*{letter}{rows["ebit_margin"]}'
            elif key == "cash_tax":
                value = "N/A" if year == 2025 else f'=Assumptions!$H${assumption_rows["OP_CASH_TAX"]}'
            elif key == "nopat":
                value = "N/A" if year == 2025 else f'={letter}{rows["ebit"]}*(1-{letter}{rows["cash_tax"]})'
            elif key == "da_pct":
                value = "N/A" if year == 2025 else f'=Assumptions!$H${assumption_rows["OP_DA_PCT_REV"]}'
            elif key == "da":
                value = "N/A" if year == 2025 else f'={letter}{rows["model_revenue"]}*{letter}{rows["da_pct"]}'
            elif key == "capex_pct":
                value = "N/A" if year == 2025 else f'=Assumptions!$H${assumption_rows["OP_CAPEX_PCT_REV"]}'
            elif key == "capex":
                value = "N/A" if year == 2025 else f'={letter}{rows["model_revenue"]}*{letter}{rows["capex_pct"]}'
            elif key == "delta_nwc":
                if year == 2025:
                    value = "N/A"
                else:
                    prev = get_column_letter(col - 1)
                    value = f'=MAX(0,{letter}{rows["model_revenue"]}-{prev}{rows["model_revenue"]})*Assumptions!$H${assumption_rows["OP_NWC_PCT_DELTA_REV"]}'
            elif key == "fcff":
                value = "N/A" if year == 2025 else f'={letter}{rows["nopat"]}+{letter}{rows["da"]}-{letter}{rows["capex"]}-{letter}{rows["delta_nwc"]}'
            elif key == "discount_period":
                value = "N/A" if year == 2025 else year - 2025.5
            elif key == "discount_factor":
                value = "N/A" if year == 2025 else f'=1/(1+$K$6)^{letter}{rows["discount_period"]}'
            elif key == "pv_fcff":
                value = "N/A" if year == 2025 else f'={letter}{rows["fcff"]}*{letter}{rows["discount_factor"]}'
            else:
                raise KeyError(key)
            ws.cell(row, col, value)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(value, str) and value.startswith("="):
                style_formula(ws.cell(row, col), "!" in value)
            else:
                ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
            if key in percent_rows:
                ws.cell(row, col).number_format = FMT_PCT
            elif key == "discount_period":
                ws.cell(row, col).number_format = "0.0x"
            elif key not in {"status"}:
                ws.cell(row, col).number_format = FMT_BN
        for col in range(2, 9):
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
        ws.row_dimensions[row].height = 25

    # Valuation summary.
    summary_headers = ["Valuation summary", "Value", "Status", "Comment"]
    for col, header in enumerate(summary_headers, 10):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 10, 13)
    summary = {
        "scenario": 5,
        "wacc": 6,
        "terminal_g": 7,
        "terminal_fcff": 8,
        "terminal_value": 9,
        "pv_terminal": 10,
        "pv_explicit": 11,
        "enterprise_value": 12,
        "tv_pct": 13,
        "gross_debt": 15,
        "cash": 16,
        "short_term_investments": 17,
        "long_term_deposits": 18,
        "lease_liability": 19,
        "other_claims": 20,
        "erablue_value": 21,
        "additional_ipo": 22,
        "equity_value": 23,
        "post_ipo_shares": 24,
        "value_per_share": 25,
        "ipo_anchor": 26,
        "market_data": 28,
        "market_conclusion": 29,
    }
    summary_data = {
        "scenario": ("Active scenario", "=Assumptions!$C$4", "Control", "Scenario switch"),
        "wacc": ("WACC", f'=Assumptions!$H${assumption_rows["VAL_WACC"]}', "Illustrative", "Nominal VND assumption"),
        "terminal_g": ("Terminal growth", f'=Assumptions!$H${assumption_rows["VAL_TERMINAL_G"]}', "Illustrative", "Must be below WACC"),
        "terminal_fcff": ("Terminal FCFF", f'=H{rows["fcff"]}*(1+K7)', "Calculated", "2030E FCFF grown by g"),
        "terminal_value": ("Terminal value", "=K8/(K6-K7)", "Calculated", "Perpetuity growth"),
        "pv_terminal": ("PV of terminal value", f'=K9*H{rows["discount_factor"]}', "Calculated", "Mid-year convention"),
        "pv_explicit": ("PV of explicit FCFF", f'=SUM(D{rows["pv_fcff"]}:H{rows["pv_fcff"]})', "Calculated", "2026E–2030E"),
        "enterprise_value": ("Enterprise value", "=K10+K11", "Calculated", "FCFF enterprise value"),
        "tv_pct": ("Terminal value / EV", "=K10/K12", "Calculated", "High percentages require caution"),
        "gross_debt": ("Gross debt", f'=Q1_Actuals!D{q1_rows["debt"]}', "Actual at 31-Mar-2026", "Short-term debt; no long-term debt line in source BS"),
        "cash": ("Cash & equivalents", f'=Q1_Actuals!D{q1_rows["cash"]}', "Actual at 31-Mar-2026", "Post balance-sheet date IPO effects are not assumed"),
        "short_term_investments": ("Short-term investments", f'=Q1_Actuals!D{q1_rows["short_term_investments"]}', "Actual at 31-Mar-2026", "Held-to-maturity investments"),
        "long_term_deposits": ("Long-term deposits", f'=Q1_Actuals!D{q1_rows["long_term_deposits"]}', "Actual at 31-Mar-2026", "Held-to-maturity investments"),
        "lease_liability": ("Lease liabilities", f'=Assumptions!$H${assumption_rows["VAL_LEASE"]}', "Illustrative placeholder", "Replace when lease convention is verified"),
        "other_claims": ("Other claims", f'=Assumptions!$H${assumption_rows["VAL_OTHER_CLAIMS"]}', "Illustrative placeholder", "No unsupported claim inserted"),
        "erablue_value": ("EraBlue stake value", f'=Assumptions!$H${assumption_rows["VAL_ERABLUE"]}', "Illustrative", "Added once; FCFF excludes JV earnings"),
        "additional_ipo": ("Additional IPO proceeds", "=IPO_Bridge!D20", "Control", "Must remain zero under post-IPO bridge policy"),
        "equity_value": ("DMX equity value", "=K12-K15+K16+K17+K18-K19-K20+K21+K22", "Calculated", "EV-to-equity bridge"),
        "post_ipo_shares": ("Post-IPO diluted shares", "=IPO_Bridge!D11", "Actual", "Million shares"),
        "value_per_share": ("DCF value per share", "=K23*1000/K24", "Illustrative", "VND/share; not a target price"),
        "ipo_anchor": ("IPO offer price anchor", "=IPO_Bridge!D5", "Actual transaction anchor", "Not a current market price"),
        "market_data": ("Same-date market price / date", '="N/A"', "Unavailable", "No price/date supplied"),
        "market_conclusion": ("Market conclusion", '="UNAVAILABLE — model output only"', "Publication guard", "No upside/downside or recommendation"),
    }
    for key, row in summary.items():
        label, value, status, comment = summary_data[key]
        ws.cell(row, 10, label)
        ws.cell(row, 11, value)
        ws.cell(row, 12, status)
        ws.cell(row, 13, comment)
        for col in range(10, 14):
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        ws.cell(row, 10).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        if isinstance(value, str) and value.startswith("="):
            style_formula(ws.cell(row, 11), "!" in value)
        if key in {"wacc", "terminal_g", "tv_pct"}:
            ws.cell(row, 11).number_format = FMT_PCT
        elif key in {"value_per_share", "ipo_anchor"}:
            ws.cell(row, 11).number_format = FMT_VND
        elif key == "post_ipo_shares":
            ws.cell(row, 11).number_format = FMT_SHARES
        elif key not in {"scenario", "market_data", "market_conclusion"}:
            ws.cell(row, 11).number_format = FMT_BN
        if key in {"market_data", "market_conclusion"}:
            for col in range(10, 14):
                ws.cell(row, col).fill = fill(PALE_YELLOW)
        ws.row_dimensions[row].height = 27

    ws.merge_cells("B28:H30")
    ws["B28"] = (
        "Key limitation: an integrated three-statement DMX forecast is not claimed here because the public case set does not contain "
        "enough segment disclosures to forecast every balance-sheet line. This transparent FCFF build therefore uses explicit cash-flow "
        "drivers and the latest actual cash/debt bridge; assumptions are editable and separately labelled."
    )
    ws["B28"].fill = fill(PALE_YELLOW)
    ws["B28"].font = Font(name="Aptos", size=9, color=NAVY, bold=True)
    ws["B28"].alignment = Alignment(wrap_text=True, vertical="center")
    return {**rows, **{f"summary_{key}": value for key, value in summary.items()}}


def build_mwg_sotp(
    wb: Workbook,
    dmx_rows: dict[str, int],
    assumption_rows: dict[str, int],
) -> dict[str, int]:
    ws = wb["MWG_SOTP"]
    setup_sheet(ws, "A5", 85)
    set_title(
        ws,
        "MWG illustrative equity-level SOTP",
        "DMX includes EraBlue once. BHX, An Khang and AVAKids sit outside the DMX perimeter. Parent-only adjustments avoid debt double count.",
        12,
    )
    set_widths(ws, {"A": 24, "B": 20, "C": 21, "D": 18, "E": 18, "F": 18, "G": 19, "H": 17, "I": 20, "J": 28, "K": 28, "L": 32})

    headers = [
        "Asset",
        "Method",
        "Metric",
        "Metric value",
        "Multiple / factor",
        "Net debt / claims",
        "Equity value 100%",
        "MWG ownership",
        "Attributable to MWG",
        "Source / model ref",
        "Double-count bucket",
        "Status / limitation",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 1, 12)

    assets = [
        (
            "DMX",
            "FCFF DCF",
            "DMX equity value",
            f'=DMX_Valuation!K{dmx_rows["summary_equity_value"]}',
            1.0,
            0.0,
            "=D5",
            f'=Assumptions!$H${assumption_rows["SOTP_DMX_OWNERSHIP"]}',
            "=G5*H5",
            "DMX_Valuation",
            "DMX_DCF_INCL_ERABLUE",
            "Illustrative; ownership is a rounded proxy",
        ),
        (
            "BHX",
            "EV / Sales",
            "Revenue",
            f'=Assumptions!$H${assumption_rows["SOTP_BHX_REVENUE"]}',
            f'=Assumptions!$H${assumption_rows["SOTP_BHX_EVS"]}',
            f'=Assumptions!$H${assumption_rows["SOTP_BHX_NET_DEBT"]}',
            "=D6*E6-F6",
            f'=Assumptions!$H${assumption_rows["SOTP_BHX_OWNERSHIP"]}',
            "=G6*H6",
            "Assumptions",
            "BHX",
            "Illustrative; requires standalone peer/DCF work",
        ),
        (
            "An Khang",
            "EV / Sales",
            "Revenue",
            f'=Assumptions!$H${assumption_rows["SOTP_AK_REVENUE"]}',
            f'=Assumptions!$H${assumption_rows["SOTP_AK_EVS"]}',
            f'=Assumptions!$H${assumption_rows["SOTP_AK_NET_DEBT"]}',
            "=D7*E7-F7",
            f'=Assumptions!$H${assumption_rows["SOTP_AK_OWNERSHIP"]}',
            "=G7*H7",
            "Assumptions",
            "AN_KHANG",
            "Illustrative; outside DMX LFL perimeter",
        ),
        (
            "AVAKids",
            "EV / Sales",
            "Revenue",
            f'=Assumptions!$H${assumption_rows["SOTP_AVA_REVENUE"]}',
            f'=Assumptions!$H${assumption_rows["SOTP_AVA_EVS"]}',
            f'=Assumptions!$H${assumption_rows["SOTP_AVA_NET_DEBT"]}',
            "=D8*E8-F8",
            f'=Assumptions!$H${assumption_rows["SOTP_AVA_OWNERSHIP"]}',
            "=G8*H8",
            "Assumptions",
            "AVAKIDS",
            "Illustrative; outside DMX LFL perimeter",
        ),
        (
            "Other investments",
            "Equity value",
            "Fixed equity value",
            f'=Assumptions!$H${assumption_rows["SOTP_OTHER_INV"]}',
            1.0,
            0.0,
            "=D9*E9-F9",
            1.0,
            "=G9*H9",
            "Assumptions",
            "OTHER_INVESTMENTS",
            "Illustrative placeholder",
        ),
        (
            "Parent non-operating assets",
            "Equity value",
            "Fixed equity value",
            f'=Assumptions!$H${assumption_rows["SOTP_PARENT_ASSETS"]}',
            1.0,
            0.0,
            "=D10*E10-F10",
            1.0,
            "=G10*H10",
            "Assumptions",
            "PARENT_NONOP_ASSETS",
            "Illustrative placeholder",
        ),
    ]
    asset_rows: dict[str, int] = {}
    for row, asset in enumerate(assets, 5):
        asset_rows[asset[0]] = row
        for col, value in enumerate(asset, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, str) and value.startswith("="):
                style_formula(ws.cell(row, col), "!" in value)
            else:
                ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        for col in (4, 6, 7, 9):
            ws.cell(row, col).number_format = FMT_BN
        ws.cell(row, 5).number_format = FMT_MULTIPLE
        ws.cell(row, 8).number_format = FMT_PCT
        ws.row_dimensions[row].height = 31

    section(ws, 13, "MWG NAV BRIDGE — ILLUSTRATIVE", 1, 6)
    summary = {
        "gav": 14,
        "parent_net_debt": 15,
        "parent_provisions": 16,
        "nav_pre_discount": 17,
        "holdco_discount": 18,
        "nav_post_discount": 19,
        "mwg_shares": 20,
        "value_per_share": 21,
        "market_data": 23,
        "market_conclusion": 24,
    }
    summary_data = {
        "gav": ("Gross attributable asset value", "=SUM(I5:I10)", "VND bn", "Sum of equity values attributable to MWG"),
        "parent_net_debt": ("Less: parent-only net debt", f'=Assumptions!$H${assumption_rows["SOTP_PARENT_NET_DEBT"]}', "VND bn", "Not consolidated net debt"),
        "parent_provisions": ("Less: parent provisions / claims", f'=Assumptions!$H${assumption_rows["SOTP_PARENT_PROVISIONS"]}', "VND bn", "Separate parent claims"),
        "nav_pre_discount": ("NAV before holdco discount", "=C14-C15-C16", "VND bn", "Presented before applying discount"),
        "holdco_discount": ("Holding-company discount", f'=Assumptions!$H${assumption_rows["SOTP_HOLDCO_DISC"]}', "%", "Sensitivity, not a target-price plug"),
        "nav_post_discount": ("NAV after holdco discount", "=C17*(1-C18)", "VND bn", "Illustrative SOTP equity value"),
        "mwg_shares": ("MWG diluted shares", f'=Assumptions!$H${assumption_rows["SOTP_MWG_SHARES"]}', "mn shares", "Unverified placeholder at this stage"),
        "value_per_share": ("Illustrative SOTP value / share", "=C19*1000/C20", "VND/share", "Not a target price"),
        "market_data": ("Same-date MWG price / date", '="N/A"', "N/A", "Market data gate not met"),
        "market_conclusion": ("Upside/downside or recommendation", '="UNAVAILABLE"', "N/A", "No market conclusion without verified price/date"),
    }
    ws.cell(13, 7, "Status")
    ws.cell(13, 8, "Interpretation")
    ws.merge_cells("H13:L13")
    for col in range(7, 13):
        ws.cell(13, col).fill = fill(NAVY_2)
        ws.cell(13, col).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        ws.cell(13, col).alignment = Alignment(horizontal="center", vertical="center")
    for key, row in summary.items():
        label, value, unit, note = summary_data[key]
        ws.cell(row, 1, label)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 3, value)
        ws.cell(row, 4, unit)
        ws.cell(row, 7, "Illustrative" if key not in {"market_data", "market_conclusion"} else "Publication guard")
        ws.cell(row, 8, note)
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
        for col in range(1, 13):
            ws.cell(row, col).fill = fill(PALE_YELLOW if key in {"market_data", "market_conclusion"} else (WHITE if row % 2 else LIGHT_GREY))
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        ws.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        if isinstance(value, str) and value.startswith("="):
            style_formula(ws.cell(row, 3), "!" in value)
        if unit == "%":
            ws.cell(row, 3).number_format = FMT_PCT
        elif unit == "VND/share":
            ws.cell(row, 3).number_format = FMT_VND
        elif unit == "mn shares":
            ws.cell(row, 3).number_format = FMT_SHARES
        else:
            ws.cell(row, 3).number_format = FMT_BN
        ws.row_dimensions[row].height = 28

    ws.merge_cells("A27:L29")
    ws["A27"] = (
        "Perimeter control — EraBlue appears only within the DMX DCF equity bridge. It is not added as a separate MWG SOTP row. "
        "BHX, An Khang and AVAKids are separately valued outside DMX. Parent-only net debt is used because subsidiary values are "
        "already equity values; consolidated MWG net debt would double count subsidiary financing."
    )
    ws["A27"].fill = fill(LIGHT_BLUE)
    ws["A27"].font = Font(name="Aptos", size=9, bold=True, color=NAVY)
    ws["A27"].alignment = Alignment(wrap_text=True, vertical="center")
    return {**asset_rows, **{f"summary_{key}": value for key, value in summary.items()}}


def build_scenarios(
    wb: Workbook,
    q1_rows: dict[str, int],
    mgmt_rows: dict[str, int],
    assumption_rows: dict[str, int],
) -> dict[str, int]:
    ws = wb["Scenarios"]
    setup_sheet(ws, "B5", 75)
    set_title(
        ws,
        "Scenario engine & valuation sensitivities",
        "Bear/Base/Bull calculations run independently from the active scenario switch. Outputs are illustrative ranges, not target prices.",
        18,
    )
    widths = {"A": 33}
    for col in range(2, 19):
        widths[get_column_letter(col)] = 14
    set_widths(ws, widths)

    summary_headers = ["Metric", "Bear", "Base", "Bull", "Status / interpretation"]
    for col, header in enumerate(summary_headers, 1):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 1, 5)

    block_specs = {
        "Bear": {"start_col": 2, "ass_col": "E"},
        "Base": {"start_col": 8, "ass_col": "F"},
        "Bull": {"start_col": 14, "ass_col": "G"},
    }
    block_rows = {
        "year": 21,
        "revenue": 22,
        "ebit_margin": 23,
        "ebit": 24,
        "nopat": 25,
        "da": 26,
        "capex": 27,
        "delta_nwc": 28,
        "fcff": 29,
        "discount_period": 30,
        "discount_factor": 31,
        "pv_fcff": 32,
        "terminal_value": 34,
        "pv_terminal": 35,
        "pv_explicit": 36,
        "enterprise_value": 37,
        "equity_value": 38,
        "value_per_share": 39,
    }
    years = list(range(2026, 2031))

    for scenario, spec in block_specs.items():
        start_col = spec["start_col"]
        end_col = start_col + 4
        ass_col = spec["ass_col"]
        ws.merge_cells(start_row=20, start_column=start_col, end_row=20, end_column=end_col)
        ws.cell(20, start_col, f"{scenario.upper()} — independent FCFF DCF")
        ws.cell(20, start_col).fill = fill(NAVY_2)
        ws.cell(20, start_col).font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        ws.cell(20, start_col).alignment = Alignment(horizontal="center")
        for col, year in enumerate(years, start_col):
            letter = get_column_letter(col)
            prev_letter = get_column_letter(col - 1)
            mgmt_col = get_column_letter(4 + year - 2026)  # D through H on Mgmt_Forecast
            ws.cell(block_rows["year"], col, f"{year}E")
            if year == 2026:
                revenue_formula = (
                    f'=Mgmt_Forecast!C{mgmt_rows["revenue"]}*'
                    f'(1+Mgmt_Forecast!{mgmt_col}{mgmt_rows["revenue_growth"]}+'
                    f'Assumptions!${ass_col}${assumption_rows["OP_REV_GROWTH_ADJ"]})'
                )
            else:
                revenue_formula = (
                    f'={prev_letter}{block_rows["revenue"]}*'
                    f'(1+Mgmt_Forecast!{mgmt_col}{mgmt_rows["revenue_growth"]}+'
                    f'Assumptions!${ass_col}${assumption_rows["OP_REV_GROWTH_ADJ"]})'
                )
            ws.cell(block_rows["revenue"], col, revenue_formula)
            if year == 2026:
                ws.cell(block_rows["ebit_margin"], col, f'=Assumptions!${ass_col}${assumption_rows["OP_EBIT_MARGIN_2026"]}')
            else:
                ws.cell(
                    block_rows["ebit_margin"],
                    col,
                    f'={prev_letter}{block_rows["ebit_margin"]}+Assumptions!${ass_col}${assumption_rows["OP_EBIT_MARGIN_STEP"]}',
                )
            ws.cell(block_rows["ebit"], col, f'={letter}{block_rows["revenue"]}*{letter}{block_rows["ebit_margin"]}')
            ws.cell(
                block_rows["nopat"],
                col,
                f'={letter}{block_rows["ebit"]}*(1-Assumptions!${ass_col}${assumption_rows["OP_CASH_TAX"]})',
            )
            ws.cell(
                block_rows["da"],
                col,
                f'={letter}{block_rows["revenue"]}*Assumptions!${ass_col}${assumption_rows["OP_DA_PCT_REV"]}',
            )
            ws.cell(
                block_rows["capex"],
                col,
                f'={letter}{block_rows["revenue"]}*Assumptions!${ass_col}${assumption_rows["OP_CAPEX_PCT_REV"]}',
            )
            prior_revenue = f'Mgmt_Forecast!C{mgmt_rows["revenue"]}' if year == 2026 else f'{prev_letter}{block_rows["revenue"]}'
            ws.cell(
                block_rows["delta_nwc"],
                col,
                f'=MAX(0,{letter}{block_rows["revenue"]}-{prior_revenue})*Assumptions!${ass_col}${assumption_rows["OP_NWC_PCT_DELTA_REV"]}',
            )
            ws.cell(
                block_rows["fcff"],
                col,
                f'={letter}{block_rows["nopat"]}+{letter}{block_rows["da"]}-{letter}{block_rows["capex"]}-{letter}{block_rows["delta_nwc"]}',
            )
            ws.cell(block_rows["discount_period"], col, year - 2025.5)
            ws.cell(
                block_rows["discount_factor"],
                col,
                f'=1/(1+Assumptions!${ass_col}${assumption_rows["VAL_WACC"]})^{letter}{block_rows["discount_period"]}',
            )
            ws.cell(block_rows["pv_fcff"], col, f'={letter}{block_rows["fcff"]}*{letter}{block_rows["discount_factor"]}')

        last = get_column_letter(end_col)
        first = get_column_letter(start_col)
        ws.cell(
            block_rows["terminal_value"],
            end_col,
            f'={last}{block_rows["fcff"]}*(1+Assumptions!${ass_col}${assumption_rows["VAL_TERMINAL_G"]})/'
            f'(Assumptions!${ass_col}${assumption_rows["VAL_WACC"]}-Assumptions!${ass_col}${assumption_rows["VAL_TERMINAL_G"]})',
        )
        ws.cell(block_rows["pv_terminal"], end_col, f'={last}{block_rows["terminal_value"]}*{last}{block_rows["discount_factor"]}')
        ws.cell(block_rows["pv_explicit"], end_col, f'=SUM({first}{block_rows["pv_fcff"]}:{last}{block_rows["pv_fcff"]})')
        ws.cell(block_rows["enterprise_value"], end_col, f'={last}{block_rows["pv_terminal"]}+{last}{block_rows["pv_explicit"]}')
        ws.cell(
            block_rows["equity_value"],
            end_col,
            f'={last}{block_rows["enterprise_value"]}'
            f'-Q1_Actuals!D{q1_rows["debt"]}+Q1_Actuals!D{q1_rows["cash"]}'
            f'+Q1_Actuals!D{q1_rows["short_term_investments"]}+Q1_Actuals!D{q1_rows["long_term_deposits"]}'
            f'-Assumptions!${ass_col}${assumption_rows["VAL_LEASE"]}-Assumptions!${ass_col}${assumption_rows["VAL_OTHER_CLAIMS"]}'
            f'+Assumptions!${ass_col}${assumption_rows["VAL_ERABLUE"]}',
        )
        ws.cell(block_rows["value_per_share"], end_col, f'={last}{block_rows["equity_value"]}*1000/IPO_Bridge!D11')

        for row, label in [
            (22, "Revenue"),
            (23, "EBIT margin"),
            (24, "EBIT"),
            (25, "NOPAT"),
            (26, "D&A"),
            (27, "Capex"),
            (28, "ΔNWC"),
            (29, "FCFF"),
            (30, "Discount period"),
            (31, "Discount factor"),
            (32, "PV FCFF"),
            (34, "Terminal value"),
            (35, "PV terminal"),
            (36, "PV explicit"),
            (37, "Enterprise value"),
            (38, "Equity value"),
            (39, "Value / share"),
        ]:
            if start_col == 2:
                ws.cell(row, 1, label)
                ws.cell(row, 1).font = Font(name="Aptos", size=9, bold=row in {29, 37, 38, 39}, color=NAVY)
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row, col)
                cell.fill = fill(WHITE if row % 2 else LIGHT_GREY)
                cell.border = Border(bottom=THIN_GREY)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    style_formula(cell, "!" in cell.value)
                if row in {23, 31}:
                    cell.number_format = FMT_PCT
                elif row == 30:
                    cell.number_format = "0.0x"
                elif row == 39:
                    cell.number_format = FMT_VND
                else:
                    cell.number_format = FMT_BN

    # Summary links and independent MWG SOTP formula.
    summary_labels = [
        "WACC",
        "Terminal growth",
        "Revenue growth adjustment",
        "2030E model revenue",
        "2030E EBIT margin",
        "2030E FCFF",
        "DMX enterprise value",
        "DMX equity value",
        "DMX DCF value / share",
        "MWG illustrative SOTP / share",
        "Holdco discount",
    ]
    summary_rows: dict[str, int] = {}
    for row, label in enumerate(summary_labels, 5):
        key = label.lower().replace(" ", "_").replace("/", "per").replace("2030e", "2030")
        summary_rows[key] = row
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        for scenario_col, (scenario, spec) in enumerate(block_specs.items(), 2):
            ass_col = spec["ass_col"]
            last_col = get_column_letter(spec["start_col"] + 4)
            if label == "WACC":
                formula = f'=Assumptions!${ass_col}${assumption_rows["VAL_WACC"]}'
            elif label == "Terminal growth":
                formula = f'=Assumptions!${ass_col}${assumption_rows["VAL_TERMINAL_G"]}'
            elif label == "Revenue growth adjustment":
                formula = f'=Assumptions!${ass_col}${assumption_rows["OP_REV_GROWTH_ADJ"]}'
            elif label == "2030E model revenue":
                formula = f'={last_col}{block_rows["revenue"]}'
            elif label == "2030E EBIT margin":
                formula = f'={last_col}{block_rows["ebit_margin"]}'
            elif label == "2030E FCFF":
                formula = f'={last_col}{block_rows["fcff"]}'
            elif label == "DMX enterprise value":
                formula = f'={last_col}{block_rows["enterprise_value"]}'
            elif label == "DMX equity value":
                formula = f'={last_col}{block_rows["equity_value"]}'
            elif label == "DMX DCF value / share":
                formula = f'={last_col}{block_rows["value_per_share"]}'
            elif label == "Holdco discount":
                formula = f'=Assumptions!${ass_col}${assumption_rows["SOTP_HOLDCO_DISC"]}'
            else:
                dmx_equity = f'{last_col}{block_rows["equity_value"]}*Assumptions!${ass_col}${assumption_rows["SOTP_DMX_OWNERSHIP"]}'
                bhx = (
                    f'(Assumptions!${ass_col}${assumption_rows["SOTP_BHX_REVENUE"]}*Assumptions!${ass_col}${assumption_rows["SOTP_BHX_EVS"]}'
                    f'-Assumptions!${ass_col}${assumption_rows["SOTP_BHX_NET_DEBT"]})*Assumptions!${ass_col}${assumption_rows["SOTP_BHX_OWNERSHIP"]}'
                )
                ak = (
                    f'(Assumptions!${ass_col}${assumption_rows["SOTP_AK_REVENUE"]}*Assumptions!${ass_col}${assumption_rows["SOTP_AK_EVS"]}'
                    f'-Assumptions!${ass_col}${assumption_rows["SOTP_AK_NET_DEBT"]})*Assumptions!${ass_col}${assumption_rows["SOTP_AK_OWNERSHIP"]}'
                )
                ava = (
                    f'(Assumptions!${ass_col}${assumption_rows["SOTP_AVA_REVENUE"]}*Assumptions!${ass_col}${assumption_rows["SOTP_AVA_EVS"]}'
                    f'-Assumptions!${ass_col}${assumption_rows["SOTP_AVA_NET_DEBT"]})*Assumptions!${ass_col}${assumption_rows["SOTP_AVA_OWNERSHIP"]}'
                )
                formula = (
                    f'=(({dmx_equity})+({bhx})+({ak})+({ava})'
                    f'+Assumptions!${ass_col}${assumption_rows["SOTP_OTHER_INV"]}+Assumptions!${ass_col}${assumption_rows["SOTP_PARENT_ASSETS"]}'
                    f'-Assumptions!${ass_col}${assumption_rows["SOTP_PARENT_NET_DEBT"]}-Assumptions!${ass_col}${assumption_rows["SOTP_PARENT_PROVISIONS"]})'
                    f'*(1-Assumptions!${ass_col}${assumption_rows["SOTP_HOLDCO_DISC"]})*1000/Assumptions!${ass_col}${assumption_rows["SOTP_MWG_SHARES"]}'
                )
            ws.cell(row, scenario_col, formula)
            ws.cell(row, scenario_col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, scenario_col).border = Border(bottom=THIN_GREY)
            style_formula(ws.cell(row, scenario_col), "!" in formula)
            if label in {"WACC", "Terminal growth", "Revenue growth adjustment", "2030E EBIT margin", "Holdco discount"}:
                ws.cell(row, scenario_col).number_format = FMT_PCT
            elif "share" in label.lower():
                ws.cell(row, scenario_col).number_format = FMT_VND
            else:
                ws.cell(row, scenario_col).number_format = FMT_BN
        ws.cell(row, 5, "Illustrative range — no market recommendation")
        ws.cell(row, 5).font = Font(name="Aptos", size=9, color=GREY)
        ws.cell(row, 5).alignment = Alignment(wrap_text=True)

    # WACC × terminal-growth sensitivity using Base operating cash flows.
    section(ws, 43, "DMX DCF VALUE / SHARE — BASE OPERATIONS, WACC × TERMINAL G", 1, 8)
    wacc_values = [0.10, 0.11, 0.12, 0.13, 0.14]
    g_values = [0.02, 0.03, 0.04, 0.05, 0.06]
    ws.cell(45, 2, "WACC ↓ / g →")
    for col, g in enumerate(g_values, 3):
        ws.cell(45, col, g)
        ws.cell(45, col).number_format = FMT_PCT
    for row, wacc in enumerate(wacc_values, 46):
        ws.cell(row, 2, wacc)
        ws.cell(row, 2).number_format = FMT_PCT
        for col in range(3, 8):
            letter = get_column_letter(col)
            formula = (
                f'=(SUMPRODUCT($H$29:$L$29,1/(1+$B{row})^{{0.5,1.5,2.5,3.5,4.5}})'
                f'+$L$29*(1+{letter}$45)/($B{row}-{letter}$45)/(1+$B{row})^4.5'
                f'-Q1_Actuals!D{q1_rows["debt"]}+Q1_Actuals!D{q1_rows["cash"]}'
                f'+Q1_Actuals!D{q1_rows["short_term_investments"]}+Q1_Actuals!D{q1_rows["long_term_deposits"]}'
                f'-Assumptions!$F${assumption_rows["VAL_LEASE"]}-Assumptions!$F${assumption_rows["VAL_OTHER_CLAIMS"]}'
                f'+Assumptions!$F${assumption_rows["VAL_ERABLUE"]})*1000/IPO_Bridge!D11'
            )
            ws.cell(row, col, formula)
            ws.cell(row, col).number_format = FMT_VND
            ws.cell(row, col).alignment = Alignment(horizontal="center")
    ws.conditional_formatting.add(
        "C46:G50",
        ColorScaleRule(start_type="min", start_color=PALE_RED, mid_type="percentile", mid_value=50, mid_color=PALE_YELLOW, end_type="max", end_color=PALE_GREEN),
    )

    # DMX equity-value multiplier × holdco discount sensitivity for the MWG SOTP.
    section(ws, 43, "MWG SOTP / SHARE — BASE INPUTS, DMX VALUE MULTIPLIER × HOLDCO DISCOUNT", 10, 16)
    discounts = [0.0, 0.10, 0.20, 0.30]
    multipliers = [0.80, 0.90, 1.00, 1.10, 1.20]
    ws.cell(45, 10, "DMX value ↓ / discount →")
    for col, disc in enumerate(discounts, 11):
        ws.cell(45, col, disc)
        ws.cell(45, col).number_format = FMT_PCT
    base_other_assets = (
        f'(Assumptions!$F${assumption_rows["SOTP_BHX_REVENUE"]}*Assumptions!$F${assumption_rows["SOTP_BHX_EVS"]}-Assumptions!$F${assumption_rows["SOTP_BHX_NET_DEBT"]})*Assumptions!$F${assumption_rows["SOTP_BHX_OWNERSHIP"]}'
        f'+(Assumptions!$F${assumption_rows["SOTP_AK_REVENUE"]}*Assumptions!$F${assumption_rows["SOTP_AK_EVS"]}-Assumptions!$F${assumption_rows["SOTP_AK_NET_DEBT"]})*Assumptions!$F${assumption_rows["SOTP_AK_OWNERSHIP"]}'
        f'+(Assumptions!$F${assumption_rows["SOTP_AVA_REVENUE"]}*Assumptions!$F${assumption_rows["SOTP_AVA_EVS"]}-Assumptions!$F${assumption_rows["SOTP_AVA_NET_DEBT"]})*Assumptions!$F${assumption_rows["SOTP_AVA_OWNERSHIP"]}'
        f'+Assumptions!$F${assumption_rows["SOTP_OTHER_INV"]}+Assumptions!$F${assumption_rows["SOTP_PARENT_ASSETS"]}'
        f'-Assumptions!$F${assumption_rows["SOTP_PARENT_NET_DEBT"]}-Assumptions!$F${assumption_rows["SOTP_PARENT_PROVISIONS"]}'
    )
    for row, multiplier in enumerate(multipliers, 46):
        ws.cell(row, 10, multiplier)
        ws.cell(row, 10).number_format = FMT_MULTIPLE
        for col in range(11, 15):
            letter = get_column_letter(col)
            formula = (
                f'=(($L$38*Assumptions!$F${assumption_rows["SOTP_DMX_OWNERSHIP"]}*$J{row})+({base_other_assets}))'
                f'*(1-{letter}$45)*1000/Assumptions!$F${assumption_rows["SOTP_MWG_SHARES"]}'
            )
            ws.cell(row, col, formula)
            ws.cell(row, col).number_format = FMT_VND
            ws.cell(row, col).alignment = Alignment(horizontal="center")
    ws.conditional_formatting.add(
        "K46:N50",
        ColorScaleRule(start_type="min", start_color=PALE_RED, mid_type="percentile", mid_value=50, mid_color=PALE_YELLOW, end_type="max", end_color=PALE_GREEN),
    )

    # Scenario chart: DMX DCF and MWG SOTP share values.
    chart = BarChart()
    chart.type = "col"
    chart.title = "Illustrative scenario value / share"
    chart.y_axis.title = "VND/share"
    chart.height = 8
    chart.width = 14
    data = Reference(ws, min_col=1, max_col=4, min_row=13, max_row=14)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    cats = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=4)
    chart.set_categories(cats)
    ws.add_chart(chart, "J4")

    return {**block_rows, **summary_rows}


def build_checks(
    wb: Workbook,
    ipo_rows: dict[str, int],
    q1_rows: dict[str, int],
    assumption_rows: dict[str, int],
    dmx_rows: dict[str, int],
    sotp_rows: dict[str, int],
) -> None:
    ws = wb["Checks"]
    setup_sheet(ws, "A5", 80)
    set_title(
        ws,
        "Model QA & publication gates",
        "Critical errors must be zero. Warnings disclose missing evidence or judgement; they are not silently overwritten.",
        10,
    )
    set_widths(ws, {"A": 31, "B": 18, "C": 13, "D": 49, "E": 15, "F": 18, "G": 22, "H": 50, "I": 25, "J": 25})
    headers = ["Check ID", "Area", "Severity", "Test", "Tolerance", "Calculated value", "Status", "Interpretation / action"]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_table_header(ws, 4, 1, 8)

    checks: list[tuple[str, str, str, str, Any, str, str, str]] = [
        (
            "CHK_IPO_SHARES",
            "IPO",
            "Critical",
            "Pre shares + allocated shares = post shares",
            0.000001,
            f'=ABS(IPO_Bridge!D{ipo_rows["Shares before IPO"]}+IPO_Bridge!D{ipo_rows["Shares successfully allocated"]}-IPO_Bridge!D{ipo_rows["Shares after IPO"]})',
            '=IF(F5<=E5,"PASS","FAIL")',
            "Share-count identity",
        ),
        (
            "CHK_IPO_PROCEEDS",
            "IPO",
            "Critical",
            "Allocated shares × offer price = officially reported VND 13,315.080bn",
            0.001,
            f'=ABS(IPO_Bridge!D{ipo_rows["Shares successfully allocated"]}*IPO_Bridge!D{ipo_rows["IPO offer price"]}/1000-13315.080)',
            '=IF(F6<=E6,"PASS","FAIL")',
            "Unit conversion: mn shares × VND/share ÷ 1,000 = VND bn",
        ),
        (
            "CHK_IPO_CANCELLED",
            "IPO",
            "Critical",
            "Offered − allocated − cancelled = 0",
            0.000001,
            f'=ABS(IPO_Bridge!D{ipo_rows["Shares offered"]}-IPO_Bridge!D{ipo_rows["Shares successfully allocated"]}-IPO_Bridge!D{ipo_rows["Shares cancelled"]})',
            '=IF(F7<=E7,"PASS","FAIL")',
            "Unallocated shares are not added to charter capital",
        ),
        (
            "CHK_IPO_VALUATION_IDENTITY",
            "IPO",
            "Critical",
            "Pre-money + gross proceeds = post-money at the same IPO price",
            0.001,
            f'=ABS(IPO_Bridge!D{ipo_rows["Pre-money equity value at IPO price"]}+IPO_Bridge!D{ipo_rows["Gross proceeds"]}-IPO_Bridge!D{ipo_rows["Post-money equity value at IPO price"]})',
            '=IF(F8<=E8,"PASS","FAIL")',
            "Transaction identity; not evidence of value creation",
        ),
        (
            "CHK_3S_BS_BALANCE",
            "Actuals",
            "Critical",
            "Assets − liabilities − equity = 0",
            0.5,
            f'=ABS(Q1_Actuals!D{q1_rows["total_assets"]}-Q1_Actuals!D{q1_rows["total_liabilities"]}-Q1_Actuals!D{q1_rows["equity"]})',
            '=IF(F9<=E9,"PASS","FAIL")',
            "31-Mar-2026 consolidated balance sheet",
        ),
        (
            "CHK_3S_CASH_ROLL",
            "Actuals",
            "Critical",
            "Opening cash + net change + FX = closing cash",
            0.5,
            f'=ABS(Q1_Actuals!D{q1_rows["cash_open"]}+Q1_Actuals!D{q1_rows["net_cash_change"]}+Q1_Actuals!D{q1_rows["fx_effect"]}-Q1_Actuals!D{q1_rows["cash_close"]})',
            '=IF(F10<=E10,"PASS","FAIL")',
            "Q1 2026 cash-flow roll-forward",
        ),
        (
            "CHK_IS_GROSS_PROFIT",
            "Actuals",
            "Critical",
            "Revenue + reported COGS − gross profit = 0",
            0.5,
            f'=ABS(Q1_Actuals!D{q1_rows["revenue"]}+Q1_Actuals!D{q1_rows["cogs"]}-Q1_Actuals!D{q1_rows["gross_profit"]})',
            '=IF(F11<=E11,"PASS","FAIL")',
            "Reported expense sign is negative",
        ),
        (
            "CHK_DCF_WACC_GT_G",
            "DCF",
            "Critical",
            "WACC must exceed terminal growth",
            0.0,
            f'=DMX_Valuation!K{dmx_rows["summary_wacc"]}-DMX_Valuation!K{dmx_rows["summary_terminal_g"]}',
            '=IF(F12>E12,"PASS","FAIL")',
            "A non-positive spread makes perpetuity value invalid",
        ),
        (
            "CHK_DCF_EV_BRIDGE",
            "DCF",
            "Critical",
            "EV = PV explicit FCFF + PV terminal value",
            0.01,
            f'=ABS(DMX_Valuation!K{dmx_rows["summary_enterprise_value"]}-DMX_Valuation!K{dmx_rows["summary_pv_explicit"]}-DMX_Valuation!K{dmx_rows["summary_pv_terminal"]})',
            '=IF(F13<=E13,"PASS","FAIL")',
            "Enterprise-value bridge",
        ),
        (
            "CHK_DCF_EQUITY_BRIDGE",
            "DCF",
            "Critical",
            "DMX equity bridge recomputes from EV and each claim/asset",
            0.01,
            (
                f'=ABS(DMX_Valuation!K{dmx_rows["summary_equity_value"]}-('
                f'DMX_Valuation!K{dmx_rows["summary_enterprise_value"]}-DMX_Valuation!K{dmx_rows["summary_gross_debt"]}'
                f'+DMX_Valuation!K{dmx_rows["summary_cash"]}+DMX_Valuation!K{dmx_rows["summary_short_term_investments"]}'
                f'+DMX_Valuation!K{dmx_rows["summary_long_term_deposits"]}-DMX_Valuation!K{dmx_rows["summary_lease_liability"]}'
                f'-DMX_Valuation!K{dmx_rows["summary_other_claims"]}+DMX_Valuation!K{dmx_rows["summary_erablue_value"]}'
                f'+DMX_Valuation!K{dmx_rows["summary_additional_ipo"]}))'
            ),
            '=IF(F14<=E14,"PASS","FAIL")',
            "No unexplained plug in the EV-to-equity bridge",
        ),
        (
            "CHK_DCF_NO_DOUBLE_COUNT_IPO",
            "DCF",
            "Critical",
            "Additional IPO proceeds in equity bridge must be zero",
            0.0,
            f'=ABS(DMX_Valuation!K{dmx_rows["summary_additional_ipo"]})',
            '=IF(F15<=E15,"PASS","FAIL")',
            "Post-IPO policy prevents proceeds being added twice",
        ),
        (
            "CHK_DCF_TV_SHARE",
            "DCF",
            "Warning",
            "Terminal-value share of EV should be reviewed if above 85%",
            0.85,
            f'=DMX_Valuation!K{dmx_rows["summary_tv_pct"]}',
            '=IF(F16<=E16,"PASS","WARN — high terminal value")',
            "Sensitivity is presented even when this check passes",
        ),
        (
            "CHK_SOTP_NO_DUPLICATE_BUCKET",
            "SOTP",
            "Critical",
            "Every double-count bucket is unique",
            0.0,
            '=COUNTA(MWG_SOTP!K5:K10)-SUMPRODUCT(1/COUNTIF(MWG_SOTP!K5:K10,MWG_SOTP!K5:K10))',
            '=IF(F17<=E17,"PASS","FAIL")',
            "EraBlue is inside the DMX DCF bucket only",
        ),
        (
            "CHK_SOTP_EQUITY_BRIDGE",
            "SOTP",
            "Critical",
            "Post-discount NAV = (GAV − parent debt − provisions) × (1 − discount)",
            0.01,
            (
                f'=ABS(MWG_SOTP!C{sotp_rows["summary_nav_post_discount"]}-('
                f'MWG_SOTP!C{sotp_rows["summary_gav"]}-MWG_SOTP!C{sotp_rows["summary_parent_net_debt"]}'
                f'-MWG_SOTP!C{sotp_rows["summary_parent_provisions"]})*(1-MWG_SOTP!C{sotp_rows["summary_holdco_discount"]}))'
            ),
            '=IF(F18<=E18,"PASS","FAIL")',
            "Equity-level SOTP bridge",
        ),
        (
            "CHK_PERIMETER_ERABLUE",
            "SOTP",
            "Critical",
            "EraBlue must not appear as a separate MWG SOTP asset row",
            0.0,
            '=COUNTIF(MWG_SOTP!A5:A10,"*EraBlue*")',
            '=IF(F19<=E19,"PASS","FAIL")',
            "EraBlue value is included once inside DMX",
        ),
        (
            "CHK_OWNERSHIP_EXACT",
            "Ownership",
            "Warning",
            "Exact post-IPO MWG ownership is required for production use",
            0.0,
            '=IF(Assumptions!C8="N/A",1,0)',
            '=IF(F20=0,"PASS","WARN — rounded proxy only")',
            "Do not infer exact shares from 'nearly 86%'",
        ),
        (
            "CHK_MARKET_PRICE_DATE",
            "Market data",
            "Warning",
            "Valuation date and same-date market price are both required",
            0.0,
            '=IF(OR(Assumptions!C5="N/A",Assumptions!C6="N/A"),1,0)',
            '=IF(F21=0,"PASS","WARN — market data absent")',
            "Expected warning; market conclusion remains blocked",
        ),
        (
            "CHK_MARKET_CONCLUSION_GUARD",
            "Publication",
            "Critical",
            "Market conclusion must remain unavailable while price/date are absent",
            0.0,
            f'=IF(AND(DMX_Valuation!K{dmx_rows["summary_market_conclusion"]}="UNAVAILABLE — model output only",MWG_SOTP!C{sotp_rows["summary_market_conclusion"]}="UNAVAILABLE"),0,1)',
            '=IF(F22=0,"PASS","FAIL")',
            "Prevents unsupported upside/downside language",
        ),
        (
            "CHK_ACTUAL_SOURCE_COMPLETE",
            "Sources",
            "Critical",
            "Every row labelled Actual has a source ID",
            0.0,
            '=COUNTIFS(Q1_Actuals!I:I,"Actual",Q1_Actuals!G:G,"")',
            '=IF(F23=0,"PASS","FAIL")',
            "Source locator is also displayed beside each actual",
        ),
        (
            "CHK_SCENARIO_LINKS",
            "Scenarios",
            "Critical",
            "Active scenario must be Bear, Base or Bull",
            0.0,
            '=IF(OR(Assumptions!C4="Bear",Assumptions!C4="Base",Assumptions!C4="Bull"),0,1)',
            '=IF(F24=0,"PASS","FAIL")',
            "Scenario selector controls active assumption formulas",
        ),
    ]

    for row, check in enumerate(checks, 5):
        check_id, area, severity, test, tolerance, value_formula, status_formula, interpretation = check
        values = [check_id, area, severity, test, tolerance, value_formula, status_formula, interpretation]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).fill = fill(WHITE if row % 2 else LIGHT_GREY)
            ws.cell(row, col).border = Border(bottom=THIN_GREY)
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, str) and value.startswith("="):
                style_formula(ws.cell(row, col), "!" in value)
            else:
                ws.cell(row, col).font = Font(name="Aptos", size=9, color=BLACK)
        ws.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        ws.cell(row, 3).font = Font(name="Aptos", size=9, bold=True, color=RED if severity == "Critical" else ORANGE)
        ws.cell(row, 5).number_format = FMT_BN_2
        ws.cell(row, 6).number_format = FMT_PCT if check_id == "CHK_DCF_TV_SHARE" else FMT_BN_2
        ws.cell(row, 7).font = Font(name="Aptos", size=9, bold=True, color=BLACK)
        ws.row_dimensions[row].height = 34

    end_row = 4 + len(checks)
    add_status_conditional_formatting(ws, f"G5:G{end_row}")
    ws.auto_filter.ref = f"A4:H{end_row}"

    # Publication dashboard. Formula cells update when Excel recalculates.
    ws["J4"] = "QA DASHBOARD"
    ws["J4"].fill = fill(NAVY)
    ws["J4"].font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    ws["I5"] = "Critical FAIL"
    ws["J5"] = f'=COUNTIFS(C5:C{end_row},"Critical",G5:G{end_row},"FAIL*")'
    ws["I6"] = "Warnings"
    ws["J6"] = f'=COUNTIF(G5:G{end_row},"WARN*")'
    ws["I7"] = "Active scenario"
    ws["J7"] = "=Assumptions!C4"
    ws["I8"] = "Market conclusion allowed"
    ws["J8"] = '=IF(AND(J5=0,J6=0,Assumptions!C5<>"N/A",Assumptions!C6<>"N/A",Assumptions!C8<>"N/A"),"YES","NO")'
    ws["I9"] = "Workbook publishable now"
    ws["J9"] = '=IF(AND(J5=0,J6=0),"YES","NO — resolve warnings")'
    for row in range(5, 10):
        ws.cell(row, 9).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        ws.cell(row, 10).font = Font(name="Aptos", size=9, bold=True, color=BLACK)
        ws.cell(row, 9).fill = fill(LIGHT_GREY)
        ws.cell(row, 10).fill = fill(PALE_YELLOW if row >= 8 else LIGHT_BLUE)
        ws.cell(row, 9).alignment = Alignment(wrap_text=True)
        ws.cell(row, 10).alignment = Alignment(wrap_text=True)


def add_workbook_navigation(wb: Workbook) -> None:
    for ws in wb.worksheets:
        if ws.title == "Cover":
            continue
        # A compact return link in the last used column does not disturb model cells.
        max_col = max(ws.max_column, 2)
        cell = ws.cell(1, max_col)
        if cell.coordinate not in ws.merged_cells:
            cell.value = "← Cover"
            cell.hyperlink = "#'Cover'!B2"
            cell.style = "Hyperlink"
            cell.alignment = Alignment(horizontal="right")


def validate_generated_workbook(path: Path) -> dict[str, int]:
    from zipfile import ZipFile

    with ZipFile(path) as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise ValueError(f"Corrupt XLSX member: {bad_member}")

    wb = load_workbook(path, read_only=False, data_only=False)
    if wb.sheetnames != SHEET_ORDER:
        raise ValueError(f"Unexpected sheet order: {wb.sheetnames}")
    formula_count = 0
    comment_count = 0
    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            raise ValueError(f"Empty worksheet: {ws.title}")
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comment_count += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if "#REF!" in cell.value or "#VALUE!" in cell.value:
                        raise ValueError(f"Broken formula token in {ws.title}!{cell.coordinate}: {cell.value}")
                    if "[" in cell.value and "]" in cell.value:
                        raise ValueError(f"External workbook link in {ws.title}!{cell.coordinate}: {cell.value}")
    if formula_count < 250:
        raise ValueError(f"Too few formulas for the requested model: {formula_count}")
    if comment_count < 20:
        raise ValueError(f"Assumption audit comments missing: {comment_count}")
    return {"formula_cells": formula_count, "comments": comment_count, "sheets": len(wb.sheetnames)}


def calculate_base_case_sanity() -> dict[str, float]:
    """Independent Python calculation used only as a build-time reasonableness check."""
    revenue = MGMT_FORECAST[2025]["revenue"]
    fcffs: list[float] = []
    for index, year in enumerate(range(2026, 2031)):
        growth = MGMT_FORECAST[year]["revenue"] / MGMT_FORECAST[year - 1]["revenue"] - 1
        prior_revenue = revenue
        revenue = prior_revenue * (1 + growth)
        ebit_margin = 0.080 + index * 0.002
        nopat = revenue * ebit_margin * (1 - 0.20)
        da = revenue * 0.010
        capex = revenue * 0.010
        delta_nwc = max(0.0, revenue - prior_revenue) * 0.030
        fcffs.append(nopat + da - capex - delta_nwc)
    wacc = 0.12
    terminal_g = 0.04
    periods = [0.5, 1.5, 2.5, 3.5, 4.5]
    pv_explicit = sum(cf / (1 + wacc) ** period for cf, period in zip(fcffs, periods))
    terminal_value = fcffs[-1] * (1 + terminal_g) / (wacc - terminal_g)
    pv_terminal = terminal_value / (1 + wacc) ** periods[-1]
    enterprise_value = pv_explicit + pv_terminal
    equity_value = (
        enterprise_value
        - Q1["short_term_debt_mar_2026"]
        + Q1["cash_mar_2026"]
        + Q1["short_term_investments_mar_2026"]
        + Q1["long_term_deposits_mar_2026"]
        + 1_500.0
    )
    value_per_share = equity_value * 1000 / 1_267.722
    tv_share = pv_terminal / enterprise_value
    if not (0 < enterprise_value < 1_000_000 and 0 < value_per_share < 1_000_000):
        raise ValueError("Base-case sanity calculation is outside broad portfolio-model bounds")
    if tv_share >= 0.95:
        raise ValueError(f"Terminal-value share is implausibly high: {tv_share:.1%}")
    return {
        "enterprise_value_bn": enterprise_value,
        "equity_value_bn": equity_value,
        "value_per_share_vnd": value_per_share,
        "terminal_value_pct": tv_share,
    }


def build_model() -> Path:
    source_messages = verify_embedded_q1_against_raw()
    wb = new_workbook()
    build_cover(wb)
    build_sources(wb)
    ipo_rows = build_ipo_bridge(wb)
    q1_rows = build_q1_actuals(wb)
    mgmt_rows = build_mgmt_forecast(wb)
    assumption_rows = build_assumptions(wb)
    dmx_rows = build_dmx_valuation(wb, q1_rows, mgmt_rows, assumption_rows)
    sotp_rows = build_mwg_sotp(wb, dmx_rows, assumption_rows)
    build_scenarios(wb, q1_rows, mgmt_rows, assumption_rows)
    build_checks(wb, ipo_rows, q1_rows, assumption_rows, dmx_rows, sotp_rows)
    add_workbook_navigation(wb)
    wb.active = wb.sheetnames.index("Cover")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    qa = validate_generated_workbook(OUTPUT_PATH)
    sanity = calculate_base_case_sanity()
    print(f"Built: {OUTPUT_PATH}")
    print(f"Workbook QA: {qa}")
    print(
        "Independent Base sanity: "
        f"EV={sanity['enterprise_value_bn']:,.1f} VND bn; "
        f"equity={sanity['equity_value_bn']:,.1f} VND bn; "
        f"value/share={sanity['value_per_share_vnd']:,.0f} VND; "
        f"TV/EV={sanity['terminal_value_pct']:.1%}"
    )
    for message in source_messages:
        print(message)
    print("Formula results recalculate when opened in Excel/LibreOffice (automatic/full calculation enabled).")
    return OUTPUT_PATH


if __name__ == "__main__":
    build_model()
